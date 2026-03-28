import sys
import subprocess
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    # 2026-03-28 15:18 修改原因：当前仓库仍然不是标准安装包结构，测试直接运行时需要显式补充项目根路径。
    # 2026-03-28 15:18 修改目的：保证老板报告工作簿测试继续能以低侵入方式直接执行，而不额外改生产入口。
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from tools.boss_report_workbook import (
    ActionCyclePoint,
    ActionImpactItem,
    BossReportBundle,
    CityContribution,
    LossActionItem,
    MonthlyTrendPoint,
    RiskCitySummary,
    ScenarioCyclePoint,
    ScenarioForecast,
    ShopCategorySummary,
    SummaryMetric,
    WarningSignal,
    WeeklyWarningPoint,
    YearSummary,
    build_boss_report_workbook,
)


def _sheet_text(workbook, sheet_name: str) -> str:
    # 2026-03-28 15:18 修改原因：这轮要验证的中文结论显著增多，重复手写遍历逻辑会让测试噪声很大。
    # 2026-03-28 15:18 修改目的：统一抽成文本提取辅助函数，让每个断言都直接表达业务合同。
    worksheet = workbook[sheet_name]
    return "\n".join(
        str(cell)
        for row in worksheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )


def _sample_bundle() -> BossReportBundle:
    # 2026-03-28 15:18 修改原因：用户已经把要求升级成“月度主轴 + 周度补充”的老板预测版报告，旧的静态样例不够用了。
    # 2026-03-28 15:18 修改目的：给测试提供稳定的时间序列、场景路径和动作拐点样例，避免依赖真实 Excel 波动。
    monthly_history = [
        MonthlyTrendPoint(period_label="2023-08", order_count=934, revenue=1282398.00, profit=37102.90),
        MonthlyTrendPoint(period_label="2023-09", order_count=939, revenue=1531944.97, profit=56014.53),
        MonthlyTrendPoint(period_label="2023-10", order_count=919, revenue=1126718.83, profit=37010.06),
        MonthlyTrendPoint(period_label="2023-11", order_count=2222, revenue=2947827.15, profit=91444.25),
        MonthlyTrendPoint(period_label="2023-12", order_count=1153, revenue=1687286.04, profit=40147.00),
    ]
    future_monthly_forecast = [
        MonthlyTrendPoint(period_label="2024-01", order_count=1180, revenue=1710000.00, profit=38200.00),
        MonthlyTrendPoint(period_label="2024-02", order_count=1165, revenue=1735000.00, profit=36000.00),
        MonthlyTrendPoint(period_label="2024-03", order_count=1190, revenue=1762000.00, profit=34700.00),
        MonthlyTrendPoint(period_label="2024-04", order_count=1215, revenue=1784000.00, profit=33800.00),
        MonthlyTrendPoint(period_label="2024-05", order_count=1240, revenue=1806000.00, profit=33100.00),
    ]
    key_drag_forecast = [
        MonthlyTrendPoint(period_label="2024-01", order_count=610, revenue=910000.00, profit=1900.00),
        MonthlyTrendPoint(period_label="2024-02", order_count=620, revenue=926000.00, profit=900.00),
        MonthlyTrendPoint(period_label="2024-03", order_count=635, revenue=941000.00, profit=-600.00),
        MonthlyTrendPoint(period_label="2024-04", order_count=648, revenue=955000.00, profit=-1800.00),
        MonthlyTrendPoint(period_label="2024-05", order_count=660, revenue=970000.00, profit=-3200.00),
    ]
    weekly_warning_signals = [
        WeeklyWarningPoint(
            period_label="2023-W50",
            signal_name="风险城市亏损扩大",
            revenue=355460.87,
            profit=12529.11,
            note="风险城市利润修复力度不足，周度失血尚未逆转。",
        ),
        WeeklyWarningPoint(
            period_label="2023-W51",
            signal_name="重点组合利润继续下探",
            revenue=350791.78,
            profit=9946.20,
            note="天猫店铺+酒店周度毛利继续被压缩，逼近失控阈值。",
        ),
        WeeklyWarningPoint(
            period_label="2023-W52",
            signal_name="单周亏损拐入负值",
            revenue=467063.62,
            profit=-6109.39,
            note="若进入下一个月仍不纠偏，月度利润率会继续恶化。",
        ),
    ]
    scenario_a_points = [
        ScenarioCyclePoint(period_label="2024-01", revenue=1710000.00, profit=38200.00, profit_rate=0.0223),
        ScenarioCyclePoint(period_label="2024-02", revenue=1735000.00, profit=36000.00, profit_rate=0.0207),
        ScenarioCyclePoint(period_label="2024-03", revenue=1762000.00, profit=34700.00, profit_rate=0.0197),
        ScenarioCyclePoint(period_label="2024-04", revenue=1784000.00, profit=33800.00, profit_rate=0.0189),
        ScenarioCyclePoint(period_label="2024-05", revenue=1806000.00, profit=33100.00, profit_rate=0.0183),
    ]
    scenario_b_points = [
        ScenarioCyclePoint(period_label="2024-01", revenue=1655000.00, profit=40500.00, profit_rate=0.0245),
        ScenarioCyclePoint(period_label="2024-02", revenue=1630000.00, profit=44000.00, profit_rate=0.0270),
        ScenarioCyclePoint(period_label="2024-03", revenue=1645000.00, profit=47700.00, profit_rate=0.0290),
        ScenarioCyclePoint(period_label="2024-04", revenue=1660000.00, profit=50500.00, profit_rate=0.0304),
        ScenarioCyclePoint(period_label="2024-05", revenue=1678000.00, profit=52100.00, profit_rate=0.0310),
    ]
    scenario_c_points = [
        ScenarioCyclePoint(period_label="2024-01", revenue=1695000.00, profit=43800.00, profit_rate=0.0258),
        ScenarioCyclePoint(period_label="2024-02", revenue=1712000.00, profit=48900.00, profit_rate=0.0286),
        ScenarioCyclePoint(period_label="2024-03", revenue=1738000.00, profit=54500.00, profit_rate=0.0314),
        ScenarioCyclePoint(period_label="2024-04", revenue=1764000.00, profit=60100.00, profit_rate=0.0341),
        ScenarioCyclePoint(period_label="2024-05", revenue=1795000.00, profit=65800.00, profit_rate=0.0367),
    ]
    action_fix_points = [
        ActionCyclePoint(period_label="2024-01", monthly_profit_uplift=12000.00, cumulative_profit_uplift=12000.00),
        ActionCyclePoint(period_label="2024-02", monthly_profit_uplift=21000.00, cumulative_profit_uplift=33000.00),
        ActionCyclePoint(period_label="2024-03", monthly_profit_uplift=31000.00, cumulative_profit_uplift=64000.00),
        ActionCyclePoint(period_label="2024-04", monthly_profit_uplift=42000.00, cumulative_profit_uplift=106000.00),
        ActionCyclePoint(period_label="2024-05", monthly_profit_uplift=52000.00, cumulative_profit_uplift=158000.00),
    ]
    action_stop_loss_points = [
        ActionCyclePoint(period_label="2024-01", monthly_profit_uplift=9000.00, cumulative_profit_uplift=9000.00),
        ActionCyclePoint(period_label="2024-02", monthly_profit_uplift=16000.00, cumulative_profit_uplift=25000.00),
        ActionCyclePoint(period_label="2024-03", monthly_profit_uplift=19000.00, cumulative_profit_uplift=44000.00),
        ActionCyclePoint(period_label="2024-04", monthly_profit_uplift=21000.00, cumulative_profit_uplift=65000.00),
        ActionCyclePoint(period_label="2024-05", monthly_profit_uplift=23000.00, cumulative_profit_uplift=88000.00),
    ]
    action_mix_points = [
        ActionCyclePoint(period_label="2024-01", monthly_profit_uplift=6000.00, cumulative_profit_uplift=6000.00),
        ActionCyclePoint(period_label="2024-02", monthly_profit_uplift=11000.00, cumulative_profit_uplift=17000.00),
        ActionCyclePoint(period_label="2024-03", monthly_profit_uplift=17000.00, cumulative_profit_uplift=34000.00),
        ActionCyclePoint(period_label="2024-04", monthly_profit_uplift=24000.00, cumulative_profit_uplift=58000.00),
        ActionCyclePoint(period_label="2024-05", monthly_profit_uplift=31000.00, cumulative_profit_uplift=89000.00),
    ]

    return BossReportBundle(
        report_title="老板决策版业绩诊断",
        source_workbook="D:/Excel测试/第3天作业-业绩诊断.xlsx",
        overall_metrics=[
            SummaryMetric(label="总订单数", value="25,545"),
            SummaryMetric(label="总销售额", value="36,302,241.87"),
            SummaryMetric(label="总毛利额", value="1,401,503.22"),
            SummaryMetric(label="整体毛利率", value="3.86%"),
        ],
        executive_points=[
            "规模增长仍在继续，但利润质量已经连续走弱。",
            "当前策略下，收入会继续放大，但利润和毛利率不会自然修复。",
            "优先级应从追规模转为止损、修结构、找拐点。",
        ],
        year_summaries=[
            YearSummary(year="2022", order_count=11470, total_sales=16421729.00, total_profit=788017.64),
            YearSummary(year="2023", order_count=14075, total_sales=19880512.87, total_profit=613485.59),
        ],
        shop_category_summaries=[
            ShopCategorySummary(shop="京东店铺", category="酒店", order_count=2494, total_sales=3670322.64, total_profit=187883.28),
            ShopCategorySummary(shop="天猫店铺", category="公寓", order_count=2538, total_sales=3486376.06, total_profit=212173.20),
            ShopCategorySummary(shop="天猫店铺", category="民宿", order_count=2166, total_sales=3225028.63, total_profit=191966.33),
            ShopCategorySummary(shop="天猫店铺", category="酒店", order_count=7472, total_sales=10718246.47, total_profit=43120.23),
            ShopCategorySummary(shop="拼多多店铺", category="酒店", order_count=2509, total_sales=3625960.99, total_profit=185243.45),
        ],
        city_contributions=[
            CityContribution(city="武汉", order_count=888, total_sales=1671317.40, total_profit=65796.84),
            CityContribution(city="上海", order_count=1272, total_sales=1626702.13, total_profit=63292.28),
            CityContribution(city="广州", order_count=1027, total_sales=1548250.02, total_profit=67503.36),
            CityContribution(city="青岛", order_count=855, total_sales=1482458.03, total_profit=33430.19),
            CityContribution(city="徐州", order_count=950, total_sales=1409776.30, total_profit=54311.22),
        ],
        low_margin_cities=[
            CityContribution(city="青岛", order_count=855, total_sales=1482458.03, total_profit=33430.19),
            CityContribution(city="苏州", order_count=602, total_sales=858598.69, total_profit=22574.97),
            CityContribution(city="天津", order_count=625, total_sales=820603.22, total_profit=27304.30),
        ],
        risk_cities=[
            RiskCitySummary(city="青岛", total_sales=653805.93, total_profit=-12591.18),
            RiskCitySummary(city="苏州", total_sales=298649.56, total_profit=-7885.53),
            RiskCitySummary(city="天津", total_sales=253415.28, total_profit=-2989.21),
        ],
        loss_actions=[
            LossActionItem(priority="P1", action="立刻复核天猫店铺酒店的定价、补贴和投放策略", owner="渠道运营"),
            LossActionItem(priority="P1", action="暂停负利润城市低价放量，先做城市止损", owner="经营分析"),
            LossActionItem(priority="P2", action="把周度毛利率预警纳入经营例会", owner="商品运营"),
        ],
        analysis_steps=[
            "先确认增长是否真的转化为利润，而不是只看规模。",
            "再拆渠道 x 品类结构，定位是谁在吞利润。",
            "随后下钻风险城市，看亏损是否在扩散。",
            "最后给出未来 5 个月预测、场景差异和动作拐点。",
        ],
        warning_signals=[
            WarningSignal(
                title="利润率持续下探预警",
                current_signal="最近月度收入仍在增长，但毛利率连续走弱。",
                consequence="如果继续当前策略，未来 5 个月利润额会继续下降，毛利率将持续下探。",
                urgency="高",
                trigger_period="2024-03",
                trend_summary="预计从 2024-03 开始进入利润修复失败区间，若不止损将继续恶化。",
            ),
            WarningSignal(
                title="大体量低质量增长预警",
                current_signal="天猫店铺 + 酒店持续放量，但利润贡献极低。",
                consequence="如果继续依赖低价和补贴换规模，收入越大，利润质量越差。",
                urgency="高",
                trigger_period="2024-02",
                trend_summary="重点拖累组合会在未来 3 个周期内跌入负利润。",
            ),
            WarningSignal(
                title="风险城市亏损扩散预警",
                current_signal="青岛、苏州、天津仍在吞噬利润。",
                consequence="如果不先止损，亏损城市会继续拖累整体利润率。",
                urgency="中高",
                trigger_period="2024-01",
                trend_summary="周度信号已经出现单周亏损，月度失血会继续放大。",
            ),
        ],
        scenario_forecasts=[
            ScenarioForecast(
                scenario_name="情景A-继续当前策略",
                revenue_change_pct=0.08,
                profit_change_pct=-0.12,
                profit_rate=0.0183,
                summary="预计收入继续增长，但利润和毛利率持续走弱。",
                turning_point_period="无自然拐点",
                monthly_projection=scenario_a_points,
                strategy_definition="维持当前放量方式，不主动止损，也不修复重点拖累组合。",
                key_actions=["维持当前投放", "不调整天猫店铺+酒店", "不收缩风险城市"],
                baseline_data_summary="之前的数据：当前月收入 1,687,286.04，月利润 40,147.00，重点拖累组合利润贡献极低。",
                conclusion_summary="收入还能继续放大，但利润不会同步改善，预测期内无自然拐点。",
                action_summary="1. 继续维持当前投放。\n2. 继续放量天猫店铺+酒店。\n3. 不处理青岛、苏州、天津等风险城市。",
                analysis_summary="1. 天猫店铺+酒店规模大但利润薄。\n2. 风险城市已经出现扩散，说明亏损不是偶发。\n3. 当前策略只会放大低质量增长。",
                data_summary="1. 当前月收入 168.73万，月利润 4.01万。\n2. 天猫店铺+酒店毛利率仅 0.40%。\n3. 拐点月份：无自然拐点。",
            ),
            ScenarioForecast(
                scenario_name="情景B-局部止损",
                revenue_change_pct=-0.01,
                profit_change_pct=0.17,
                profit_rate=0.0310,
                summary="短期收入承压，但利润会先止跌并逐步回升。",
                turning_point_period="2024-02",
                monthly_projection=scenario_b_points,
                strategy_definition="先做止损，把确定性亏损城市和低价订单压住。",
                key_actions=["收缩负利润城市低价订单", "压缩高风险投放", "优先守利润"],
                baseline_data_summary="之前的数据：风险城市已出现持续失血，周度信号出现单周亏损。",
                conclusion_summary="先止血再修复，短期牺牲一点规模，可以更快把利润拉回正轨。",
                action_summary="1. 第一批止损城市先控青岛、苏州、天津。\n2. 暂停这些城市在天猫店铺+酒店的低价放量单。\n3. 第二批跟进重庆、济南，投放从追规模切到守利润。",
                analysis_summary="1. 风险城市亏损已经不是个别点状问题。\n2. 先切确定性亏损单，比全面收缩更高效。\n3. 只要先止住青岛、苏州、天津，利润修复就会提速。",
                data_summary="1. 前3个风险城市合计失血 2.35万，前5个城市失血 2.66万。\n2. 策略利润变化 +17.00%。\n3. 拐点月份：2024-02。",
            ),
            ScenarioForecast(
                scenario_name="情景C-结构优化",
                revenue_change_pct=0.05,
                profit_change_pct=0.28,
                profit_rate=0.0367,
                summary="收入和利润可以同步改善，是更优路径。",
                turning_point_period="2024-02",
                monthly_projection=scenario_c_points,
                strategy_definition="在止损基础上同步修复重点拖累组合，并提升高毛利结构占比。",
                key_actions=["修复天猫店铺+酒店毛利率", "提升高毛利品类占比", "止损与结构优化并行"],
                baseline_data_summary="之前的数据：重点拖累组合销售占比高，但毛利贡献异常低。",
                conclusion_summary="不是简单砍业务，而是把低质量收入搬到高质量结构里，让收入和利润一起修复。",
                action_summary="1. 修复天猫店铺+酒店的定价、补贴和投放门槛。\n2. 新增投放优先导向天猫店铺公寓和民宿。\n3. 止损与结构优化并行推进。",
                analysis_summary="1. 天猫店铺+酒店不是不能做，而是当前价格和流量结构出了问题。\n2. 同店铺下公寓和民宿毛利率明显更高，说明结构替换有空间。\n3. 每转移100万销售额到更优结构，理论上可多贡献约5.5万毛利。",
                data_summary="1. 天猫店铺+酒店销售 1071.82万，毛利率 0.40%。\n2. 天猫店铺+公寓 6.09%，天猫店铺+民宿 5.95%。\n3. 拐点月份：2024-02。",
            ),
        ],
        action_impacts=[
            ActionImpactItem(
                priority="P1",
                action_name="修复天猫店铺酒店毛利率",
                improve_target="直接修复当前最大拖累组合。",
                expected_profit_uplift=158000.00,
                expected_revenue_impact="短期收入持平到小幅回落，但利润质量显著改善。",
                if_not_do="收入继续放大，但利润将在 3 个周期内转负。",
                start_period="2024-01",
                effect_period="2024-02",
                turning_point_period="2024-03",
                monthly_projection=action_fix_points,
            ),
            ActionImpactItem(
                priority="P1",
                action_name="收缩负利润城市低价订单",
                improve_target="先止住确定性亏损。",
                expected_profit_uplift=88000.00,
                expected_revenue_impact="短期收入承压，但亏损会快速收敛。",
                if_not_do="风险城市会继续扩大经营失血。",
                start_period="2024-01",
                effect_period="2024-01",
                turning_point_period="2024-02",
                monthly_projection=action_stop_loss_points,
            ),
            ActionImpactItem(
                priority="P2",
                action_name="提升高毛利品类占比",
                improve_target="把增长重新导向高质量结构。",
                expected_profit_uplift=89000.00,
                expected_revenue_impact="收入整体平稳，利润弹性更好。",
                if_not_do="结构修复会中断，利润率难以持续改善。",
                start_period="2024-02",
                effect_period="2024-03",
                turning_point_period="2024-04",
                monthly_projection=action_mix_points,
            ),
        ],
        diagnosis_summary=[
            "当前策略继续下去，收入会被天猫店铺 + 酒店继续放大，但利润不会同步增长。",
            "预计未来 5 个月整体利润额继续承压，重点拖累组合将在第 3 个周期进入负利润。",
        ],
        monthly_history=monthly_history,
        future_monthly_forecast=future_monthly_forecast,
        key_drag_monthly_forecast=key_drag_forecast,
        weekly_warning_points=weekly_warning_signals,
    )


def test_build_boss_report_workbook_adds_monthly_forecast_to_summary_and_diagnosis(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    summary_text = _sheet_text(workbook, "01_结论总览")
    diagnosis_text = _sheet_text(workbook, "03_核心诊断")

    # 2026-03-28 15:18 修改原因：用户明确要求第一页和核心诊断必须从静态结论升级到未来 3-5 个周期判断。
    # 2026-03-28 15:18 修改目的：锁定“未来5个月经营预测”“诊断结论”“重点拖累未来路径”这些老板级阅读锚点。
    assert "未来5个月经营预测" in summary_text
    assert "预计后果趋势" in summary_text
    assert "2024-01" in summary_text
    assert "一句话结论" in summary_text
    assert "问题不是收入不增长，而是增长没有转化成利润" in summary_text
    assert "诊断结论" in diagnosis_text
    assert "未来5个月预测" in diagnosis_text
    assert "天猫店铺 + 酒店" in diagnosis_text
    assert "继续放大，但利润不会同步增长" in diagnosis_text
    assert "钱漏在哪" in diagnosis_text
    assert "低质量增长" in diagnosis_text


def test_build_boss_report_workbook_adds_time_warning_and_scenario_turning_points(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    warning_text = _sheet_text(workbook, "04_经营预警")
    scenario_text = _sheet_text(workbook, "05_未来场景预测")

    # 2026-03-28 15:18 修改原因：用户指出经营预警和未来场景页都缺少时间趋势、预测周期和拐点。
    # 2026-03-28 15:18 修改目的：锁定“时间趋势预警”“周度补充信号”“预计触发月份”“拐点月份”等新合同。
    assert "时间趋势预警" in warning_text
    assert "周度补充信号" in warning_text
    assert "预计触发月份" in warning_text
    assert "2023-W52" in warning_text
    assert "如果不处理" in warning_text
    assert "后果是什么" in warning_text
    assert "情景A-继续当前策略" in scenario_text
    assert "情景B-局部止损" in scenario_text
    assert "情景C-结构优化" in scenario_text
    assert "未来5个月情景测算" in scenario_text
    assert "拐点月份" in scenario_text
    assert "无自然拐点" in scenario_text
    assert "老板可选路径" in scenario_text
    assert "策略结论" in scenario_text
    assert "策略动作" in scenario_text
    assert "策略分析" in scenario_text
    assert "策略数据" in scenario_text
    assert "天猫店铺+酒店" in scenario_text
    assert "青岛、苏州、天津" in scenario_text
    assert "每转移100万销售额" in scenario_text
    assert "预测数据" in scenario_text
    assert "拐点标记" in scenario_text

    scenario_sheet = workbook["05_未来场景预测"]
    turning_point_cell = None
    turning_flag_cell = None
    for row in scenario_sheet.iter_rows():
        values = [cell.value for cell in row]
        if len(values) >= 7 and values[0] == "情景B-局部止损" and values[1] == "2024-02":
            turning_point_cell = row[1]
            turning_flag_cell = row[6]
            break

    # 2026-03-28 16:10 修改原因：用户明确要求未来场景页不只是写拐点月份，还要在表里把拐点标红。
    # 2026-03-28 16:10 修改目的：锁定“情景B 的 2024-02 为拐点，且拐点标记使用红色强调”这一可视化合同。
    assert turning_point_cell is not None
    assert turning_flag_cell is not None
    assert turning_flag_cell.value == "拐点"
    assert turning_flag_cell.font.color is not None


def test_build_boss_report_workbook_adds_multi_turning_points_to_warning_and_scenarios(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    warning_text = _sheet_text(workbook, "04_经营预警")
    scenario_text = _sheet_text(workbook, "05_未来场景预测")

    # 2026-03-29 02:16 修改原因：用户最新批准方案A，要求拐点不能只看单一月份，而要把利润、毛利率、灯色、动作和老板主拐点全部纳入。
    # 2026-03-29 02:16 修改目的：先锁定 04/05 页新的多拐点合同，再按 TDD 最小实现“预警演化 -> 情景分化 -> 老板主拐点”主线。
    assert "预警时间轴" in warning_text
    assert "预计灯色变化" in warning_text
    assert "预计主拐点" in warning_text
    assert "风险灯色" in warning_text
    assert "老板主拐点" in scenario_text
    assert "利润拐点" in scenario_text
    assert "毛利率拐点" in scenario_text
    assert "灯色拐点" in scenario_text
    assert "动作拐点" in scenario_text
    assert "动作状态" in scenario_text
    assert "主拐点" in scenario_text


def test_build_boss_report_workbook_adds_action_cycles_and_more_supporting_charts(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    action_text = _sheet_text(workbook, "06_动作-改善测算")
    total_chart_count = sum(len(sheet._charts) for sheet in workbook.worksheets)

    # 2026-03-28 15:18 修改原因：动作页现在不能只写“做什么”，还必须写“何时开始见效、何时到达拐点”。
    # 2026-03-28 15:18 修改目的：强制实现动作周期、拐点和累计改善路径，同时抬高图表下限以覆盖新增时间序列图。
    assert "启动周期" in action_text
    assert "见效周期" in action_text
    assert "预计拐点" in action_text
    assert "累计改善毛利" in action_text
    assert "2024-03" in action_text
    assert "先止损，再修复，再优化" in action_text
    assert "建议老板本月拍板" in action_text
    assert total_chart_count >= 10


def test_build_boss_report_workbook_adds_appendix_model_explanation(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    appendix_text = _sheet_text(workbook, "08_附录-图表与明细")

    # 2026-03-28 22:08 修改原因：用户要求附录不能只放明细，还要明确解释未来场景与拐点是怎么推演出来的。
    # 2026-03-28 22:08 修改目的：锁定“算法名称、推演步骤、拐点规则、非黑盒说明”这些老板追问时必须能落地复核的文本合同。
    assert "情景经营轨迹模型" in appendix_text
    assert "加权移动平均" in appendix_text
    assert "动作改善斜率" in appendix_text
    assert "盈亏平衡穿越" in appendix_text
    assert "利润连续2期改善" in appendix_text
    assert "毛利率连续2期改善" in appendix_text
    assert "不是机器学习黑盒预测" in appendix_text


def test_build_boss_report_workbook_turns_city_sheet_into_loss_control_matrix(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    city_text = _sheet_text(workbook, "07_客户贡献拆解")

    # 2026-03-28 22:26 修改原因：用户已批准把客户贡献拆解页重构成“止损对象优先级矩阵”，不再停留在 Top 城市和低毛利城市展示。
    # 2026-03-28 22:26 修改目的：锁定执行层合同，确保 07 页直接回答先止损谁、为什么、怎么动、多久复盘。
    assert "止损对象优先级矩阵" in city_text
    assert "P1 立即止损" in city_text
    assert "P2 限制放量" in city_text
    assert "P3 结构修复" in city_text
    assert "P4 持续观察" in city_text
    assert "止损对象" in city_text
    assert "优先级" in city_text
    assert "动作" in city_text
    assert "原因" in city_text
    assert "预计改善" in city_text
    assert "观察周期" in city_text
    assert "天猫店铺+酒店" in city_text
    assert "青岛" in city_text


def test_build_boss_report_workbook_adds_execution_board_and_appendix_tracker(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    city_text = _sheet_text(workbook, "07_客户贡献拆解")
    appendix_text = _sheet_text(workbook, "08_附录-图表与明细")

    # 2026-03-29 00:18 修改原因：用户批准把 07 页继续升级为“老板页 + 执行附表联动版”，不满足于只给优先级矩阵。
    # 2026-03-29 00:18 修改目的：先锁定老板拍板口径和附录执行跟踪合同，确保后续改造直接回答“谁负责、何时完成、如何验收”。
    assert "老板拍板提示" in city_text
    assert "结论" in city_text
    assert "关键动作" in city_text
    assert "负责人" in city_text
    assert "时间表" in city_text
    assert "验收指标" in city_text
    assert "执行跟踪附表" in appendix_text
    assert "第一阶段目标" in appendix_text
    assert "风险提示" in appendix_text
    assert "复盘周期" in appendix_text


def test_build_boss_report_workbook_adds_weekly_rag_board(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    city_text = _sheet_text(workbook, "07_客户贡献拆解")
    appendix_text = _sheet_text(workbook, "08_附录-图表与明细")

    # 2026-03-29 00:44 修改原因：用户要求把执行附表进一步升级成“红黄绿周会版”，让老板看灯、团队看表。
    # 2026-03-29 00:44 修改目的：先锁定状态灯、本周判断、协同人、本周动作、下周动作与截止时间等周会合同，再做实现。
    assert "状态灯" in city_text
    assert "本周判断" in city_text
    assert "下次复盘时间" in city_text
    assert "红灯" in city_text
    assert "黄灯" in city_text
    assert "绿灯" in city_text
    assert "青岛" in city_text
    assert "天猫店铺+酒店" in city_text
    assert "武汉" in city_text
    assert "协同人" in appendix_text
    assert "本周动作" in appendix_text
    assert "下周动作" in appendix_text
    assert "截止时间" in appendix_text


def test_build_boss_report_workbook_adds_dynamic_rag_reasoning(tmp_path: Path) -> None:
    output_path = tmp_path / "boss_report_monthly.xlsx"

    build_boss_report_workbook(_sample_bundle(), output_path)

    workbook = load_workbook(output_path)
    city_text = _sheet_text(workbook, "07_客户贡献拆解")
    appendix_text = _sheet_text(workbook, "08_附录-图表与明细")

    # 2026-03-29 01:03 修改原因：用户要求把红黄绿继续从静态映射升级成动态切灯模型，老板必须看到灯色变化逻辑。
    # 2026-03-29 01:03 修改目的：先锁定“上期灯色 / 本期灯色 / 变灯原因”合同，再实现动态切灯规则。
    assert "上期灯色" in city_text
    assert "本期灯色" in city_text
    assert "变灯原因" in city_text
    assert "上期灯色" in appendix_text
    assert "本期灯色" in appendix_text
    assert "变灯原因" in appendix_text
    assert "青岛" in city_text
    assert "红灯" in city_text
    assert "天猫店铺+酒店" in city_text
    assert "黄灯" in city_text
    assert "武汉" in city_text
    assert "绿灯" in city_text


def test_boss_report_workbook_cli_supports_direct_script_help() -> None:
    script_path = PROJECT_ROOT / "tools" / "boss_report_workbook.py"

    # 2026-03-28 21:47 修改原因：真实交付时直接运行脚本会命中包导入路径问题，用户后续用本地工具链时会被这个入口绊住。
    # 2026-03-28 21:47 修改目的：先锁定“python tools\\boss_report_workbook.py --help 能正常执行”的命令行合同，再修复入口路径。
    result = subprocess.run(
        [sys.executable, str(script_path), "--help"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
