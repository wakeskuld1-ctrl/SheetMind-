import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


# 2026-03-22 13:36 CST: 修改原因和目的：先用测试锁定“只改目标 sheet、保留表头、生成有季节性的假数据”这三个核心行为，避免直接处理真实文件时误伤结构。
try:
    from tools.excel_desensitize import build_fake_rows, rewrite_workbook_copy
except ModuleNotFoundError:  # pragma: no cover - 这里保留失败态，符合先红后绿的 TDD 节奏。
    build_fake_rows = None
    rewrite_workbook_copy = None


class ExcelDesensitizeTests(unittest.TestCase):
    # 2026-03-22 13:36 CST: 修改原因和目的：统一临时目录生命周期，确保测试生成的工作簿不会污染真实客户目录。
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    # 2026-03-22 13:36 CST: 修改原因和目的：显式回收临时目录，避免测试结束后遗留中间文件。
    def tearDown(self):
        self.temp_dir.cleanup()

    # 2026-03-22 13:36 CST: 修改原因和目的：构造一个接近真实“数据处理器”的最小样本，用来验证只改“客户信息”sheet。
    def _create_processor_workbook(self, path: Path) -> None:
        wb = Workbook()
        home = wb.active
        home.title = "首页"
        home.append(["项目", "值", "说明"])
        home.append(["处理状态", "DONE", "不应被脱敏逻辑改写"])

        customer = wb.create_sheet("客户信息")
        customer.append(["客户名称", "标准客户名称", "客户分类", "A角", "B角", "是否有效", "联系电话", "保费"])
        customer.append(["新疆真实客户A", "新疆真实客户A", "重点客户", "张三", "李四", "是", "13800001111", 98000])
        customer.append(["新疆真实客户B", "新疆真实客户B", "普通客户", "王五", "赵六", "是", "13900002222", 12000])
        wb.save(path)

    # 2026-03-22 13:36 CST: 修改原因和目的：构造一个普通业务台账样本，用来验证所有目标sheet都能保留表头并替换数据区。
    def _create_ledger_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "旅责险"
        ws.append(["客户名称", "保险公司", "业务人员", "会计月度", "业务收入（人民币）", "险种"])
        ws.append(["真实旅行社", "真实保险公司", "真实人员", "2026-01", 3000, "旅责险"])
        ws.append(["真实旅行社2", "真实保险公司", "真实人员", "2026-07", 8000, "旅责险"])

        other = wb.create_sheet("团意险")
        other.append(["客户名称", "保险公司", "业务人员", "会计月度", "业务收入（人民币）", "险种"])
        other.append(["真实团险客户", "真实保险公司", "真实人员", "2026-02", 1500, "团体意外险"])
        wb.save(path)

    # 2026-03-22 13:36 CST: 修改原因和目的：先验证模块存在，再在同一个测试里锁定“只改指定sheet”的业务边界。
    def test_rewrite_only_selected_sheet_for_processor_workbook(self):
        self.assertIsNotNone(rewrite_workbook_copy, "rewrite_workbook_copy 应先通过测试驱动实现")

        source = self.root / "数据处理器.xlsm"
        output_dir = self.root / "输出"
        self._create_processor_workbook(source)

        result = rewrite_workbook_copy(
            source_path=source,
            output_dir=output_dir,
            target_sheet_names={"客户信息"},
            workbook_kind="processor",
        )

        self.assertTrue(result.exists(), "应生成脱敏副本")
        wb = load_workbook(result)
        self.assertEqual(wb["首页"]["A2"].value, "处理状态")
        self.assertEqual(wb["首页"]["B2"].value, "DONE")
        self.assertEqual(wb["客户信息"]["A1"].value, "客户名称")
        self.assertNotEqual(wb["客户信息"]["A2"].value, "新疆真实客户A")
        self.assertNotEqual(wb["客户信息"]["G2"].value, "13800001111")
        self.assertEqual(wb["客户信息"].max_row, 3)
        wb.close()

    # 2026-03-22 13:36 CST: 修改原因和目的：验证普通台账sheet会整体替换数据，但表头和行数保持稳定，便于真实文件直接复用。
    def test_rewrite_all_target_sheets_and_preserve_headers(self):
        self.assertIsNotNone(rewrite_workbook_copy, "rewrite_workbook_copy 应先通过测试驱动实现")

        source = self.root / "2026文旅体台账.xlsx"
        output_dir = self.root / "输出"
        self._create_ledger_workbook(source)

        result = rewrite_workbook_copy(
            source_path=source,
            output_dir=output_dir,
            target_sheet_names={"旅责险", "团意险"},
            workbook_kind="ledger_2026",
        )

        self.assertTrue(result.exists(), "应生成台账脱敏副本")
        wb = load_workbook(result)
        self.assertEqual(wb["旅责险"]["A1"].value, "客户名称")
        self.assertEqual(wb["团意险"]["D1"].value, "会计月度")
        self.assertEqual(wb["旅责险"].max_row, 3)
        self.assertNotEqual(wb["旅责险"]["A2"].value, "真实旅行社")
        self.assertNotEqual(wb["团意险"]["B2"].value, "真实保险公司")
        wb.close()

    # 2026-03-22 13:36 CST: 修改原因和目的：用纯函数测试锁定“淡季低、旺季高”的生成规则，防止假数据虽然随机但不符合保险业务节奏。
    def test_generated_rows_show_low_and_high_seasons(self):
        self.assertIsNotNone(build_fake_rows, "build_fake_rows 应先通过测试驱动实现")

        headers = ["客户名称", "保险公司", "业务人员", "会计月度", "业务收入（人民币）", "险种"]
        rows = build_fake_rows(headers=headers, row_count=12, workbook_kind="ledger_2026", sheet_name="旅责险")

        self.assertEqual(len(rows), 12)
        month_to_income = {row[3]: row[4] for row in rows}
        self.assertIn("2026-02", month_to_income)
        self.assertIn("2026-07", month_to_income)
        self.assertGreater(month_to_income["2026-07"], month_to_income["2026-02"])
        self.assertTrue(all(row[5] for row in rows), "险种字段应有合理假值")

    # 2026-03-22 15:26 CST: 修改原因和目的：锁定“产险主题工作簿需要同时改文件名、sheet名、首行列头与产险产品语义”，避免只改外壳不改业务主题。
    def test_property_theme_rewrites_sheet_name_and_headers(self):
        self.assertIsNotNone(rewrite_workbook_copy, "rewrite_workbook_copy 应先通过测试驱动实现")

        source = self.root / "source.xlsx"
        output_dir = self.root / "输出"
        wb = Workbook()
        ws = wb.active
        ws.title = "文旅产业链+旅责险+责任保险"
        ws.append(["经营单位", "业务部门", "客户名称", "险种"])
        ws.append(["真实单位", "真实部门", "真实客户", "真实险种"])
        wb.save(source)

        result = rewrite_workbook_copy(
            source_path=source,
            output_dir=output_dir,
            target_sheet_names=None,
            workbook_kind="property_2025",
            output_name="澄岳保险集团-产险事业群-2025经营收入总台账.xlsx",
        )

        out = load_workbook(result)
        self.assertIn("综合产险业务台账", out.sheetnames)
        themed = out["综合产险业务台账"]
        self.assertEqual(themed["A1"].value, "经营机构")
        self.assertEqual(themed["B1"].value, "所属团队")
        self.assertEqual(themed["C1"].value, "投保主体")
        self.assertIn("险", themed["D2"].value)
        self.assertTrue(result.name.startswith("澄岳保险集团-产险事业群"))
        out.close()

    # 2026-03-22 15:26 CST: 修改原因和目的：锁定“寿险主题”必须输出寿险语义产品，防止仍混出企财险、责任险等产险词。
    def test_life_theme_generates_life_products(self):
        self.assertIsNotNone(build_fake_rows, "build_fake_rows 应先通过测试驱动实现")

        headers = ["投保主体", "承保机构", "会计期间", "标准产品分类"]
        rows = build_fake_rows(headers=headers, row_count=6, workbook_kind="life_2026", sheet_name="个险长期险台账")

        self.assertEqual(len(rows), 6)
        products = {row[3] for row in rows}
        self.assertTrue(any("寿" in product or "年金" in product or "重疾" in product for product in products))
        self.assertFalse(any("企财" in product or "责任" in product or "车" in product for product in products))

    # 2026-03-22 15:26 CST: 修改原因和目的：锁定“中台工作簿”首行标题和sheet命名都要中性化，避免残留客户信息等原始命名。
    def test_processor_theme_renames_sheet_and_first_row(self):
        self.assertIsNotNone(rewrite_workbook_copy, "rewrite_workbook_copy 应先通过测试驱动实现")

        source = self.root / "processor.xlsm"
        output_dir = self.root / "输出"
        self._create_processor_workbook(source)

        result = rewrite_workbook_copy(
            source_path=source,
            output_dir=output_dir,
            target_sheet_names=None,
            workbook_kind="ops_center",
            output_name="澄岳保险集团-经营管理中台-业务数据处理器.xlsm",
        )

        out = load_workbook(result)
        self.assertIn("客户主数据", out.sheetnames)
        self.assertIn("中台首页", out.sheetnames)
        themed_customer = out["客户主数据"]
        themed_home = out["中台首页"]
        self.assertEqual(themed_customer["A1"].value, "主体名称")
        self.assertEqual(themed_customer["D1"].value, "主责经理")
        self.assertEqual(themed_home["A1"].value, "配置项")
        self.assertTrue(result.name.startswith("澄岳保险集团-经营管理中台"))
        out.close()

    # 2026-03-22 15:53 CST: 修改原因和目的：锁定地域脱敏必须全国化，避免再次生成哈密、吐鲁番等强指向性城市名。
    def test_generated_locations_use_national_city_mix_instead_of_xinjiang_only(self):
        self.assertIsNotNone(build_fake_rows, "build_fake_rows 应先通过测试驱动实现")

        headers = ["经营机构", "投保主体"]
        rows = build_fake_rows(headers=headers, row_count=16, workbook_kind="property_2025", sheet_name="综合产险业务台账")

        joined = " ".join(str(cell) for row in rows for cell in row)
        self.assertFalse(any(city in joined for city in ["哈密", "吐鲁番", "喀什", "阿勒泰", "石河子"]))
        self.assertTrue(any(city in joined for city in ["北京", "上海", "广州", "成都", "武汉", "杭州"]))


if __name__ == "__main__":
    unittest.main()
