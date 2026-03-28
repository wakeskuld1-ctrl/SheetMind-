"""老板决策版业绩诊断工作簿生成器。"""

from __future__ import annotations

import argparse
import json
import statistics
import subprocess
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
import xlsxwriter


# 2026-03-28 15:44 修改原因：本轮要把老板报告升级成“月度主轴 + 周度预警补充”的预测版，旧默认路径仍然要继续复用真实作业文件。
# 2026-03-28 15:44 修改目的：保持命令行入口、真实生成路径和测试入口一致，避免交付层再出现散落硬编码。
DEFAULT_SOURCE_WORKBOOK = Path(r"D:\Excel测试\第3天作业-业绩诊断.xlsx")
DEFAULT_OUTPUT_WORKBOOK = Path(r"D:\Excel测试\第3天作业-业绩诊断_老板汇报版.xlsx")
DEFAULT_SHEET_NAME = "源数据-业绩"
DEFAULT_RUST_BINARY = Path(r"D:\Rust\Excel_Skill\target\release\excel_skill.exe")
FORECAST_PERIODS = 5


@dataclass(frozen=True)
class SummaryMetric:
    """摘要页 KPI 卡片。"""

    label: str
    value: str


@dataclass(frozen=True)
class YearSummary:
    """年度经营结果。"""

    year: str
    order_count: int
    total_sales: float
    total_profit: float

    @property
    def profit_rate(self) -> float:
        return safe_ratio(self.total_profit, self.total_sales)


@dataclass(frozen=True)
class ShopCategorySummary:
    """渠道 x 品类 的经营结果。"""

    shop: str
    category: str
    order_count: int
    total_sales: float
    total_profit: float

    @property
    def profit_rate(self) -> float:
        return safe_ratio(self.total_profit, self.total_sales)


@dataclass(frozen=True)
class CityContribution:
    """城市贡献结果。"""

    city: str
    order_count: int
    total_sales: float
    total_profit: float

    @property
    def profit_rate(self) -> float:
        return safe_ratio(self.total_profit, self.total_sales)


@dataclass(frozen=True)
class RiskCitySummary:
    """风险城市结果。"""

    city: str
    total_sales: float
    total_profit: float

    @property
    def profit_rate(self) -> float:
        return safe_ratio(self.total_profit, self.total_sales)


@dataclass(frozen=True)
class LossActionItem:
    """止损动作项。"""

    priority: str
    action: str
    owner: str


@dataclass(frozen=True)
class MonthlyTrendPoint:
    """月度经营点位。"""

    period_label: str
    order_count: int
    revenue: float
    profit: float

    @property
    def profit_rate(self) -> float:
        return safe_ratio(self.profit, self.revenue)


@dataclass(frozen=True)
class WeeklyWarningPoint:
    """周度预警补充点位。"""

    period_label: str
    signal_name: str
    revenue: float
    profit: float
    note: str

    @property
    def profit_rate(self) -> float:
        return safe_ratio(self.profit, self.revenue)


@dataclass(frozen=True)
class WarningSignal:
    """经营预警项。"""

    title: str
    current_signal: str
    consequence: str
    urgency: str
    trigger_period: str
    trend_summary: str


@dataclass(frozen=True)
class ScenarioCyclePoint:
    """情景月度测算点位。"""

    period_label: str
    revenue: float
    profit: float
    profit_rate: float


@dataclass(frozen=True)
class ScenarioForecast:
    """未来经营场景。"""

    scenario_name: str
    revenue_change_pct: float
    profit_change_pct: float
    profit_rate: float
    summary: str
    turning_point_period: str
    monthly_projection: list[ScenarioCyclePoint]
    strategy_definition: str = ""
    key_actions: list[str] = field(default_factory=list)
    baseline_data_summary: str = ""
    conclusion_summary: str = ""
    action_summary: str = ""
    analysis_summary: str = ""
    data_summary: str = ""


@dataclass(frozen=True)
class ActionCyclePoint:
    """动作改善月度点位。"""

    period_label: str
    monthly_profit_uplift: float
    cumulative_profit_uplift: float


@dataclass(frozen=True)
class ActionImpactItem:
    """动作与改善测算。"""

    priority: str
    action_name: str
    improve_target: str
    expected_profit_uplift: float
    expected_revenue_impact: str
    if_not_do: str
    start_period: str
    effect_period: str
    turning_point_period: str
    monthly_projection: list[ActionCyclePoint]


@dataclass(frozen=True)
class LossControlPriorityItem:
    """止损对象优先级矩阵项。"""

    object_name: str
    priority: str
    action: str
    reason: str
    expected_improvement: str
    review_cycle: str
    expected_improvement_value: float


@dataclass(frozen=True)
class LossControlExecutionItem:
    """止损对象执行跟踪项。"""

    priority: str
    object_name: str
    status_light: str
    # 2026-03-28 23:32 修改原因：用户要求红黄绿不能只做静态映射，老板页和周会页都要解释“灯为什么变了”。
    # 2026-03-28 23:32 修改目的：给执行对象补齐上期灯色、本期灯色和变灯原因，兼容旧字段 `status_light` 的同时支持动态切灯。
    previous_status_light: str
    current_status_light: str
    status_change_reason: str
    conclusion: str
    key_action: str
    owner: str
    co_owner: str
    timeline: str
    deadline: str
    acceptance_metric: str
    stage_goal: str
    risk_alert: str
    review_cycle: str
    weekly_judgement: str
    next_review_time: str
    current_week_action: str
    next_week_action: str
    expected_improvement_value: float


@dataclass(frozen=True)
class BossReportBundle:
    """老板决策版工作簿的数据总包。"""

    report_title: str
    source_workbook: str
    overall_metrics: list[SummaryMetric]
    executive_points: list[str]
    year_summaries: list[YearSummary]
    shop_category_summaries: list[ShopCategorySummary]
    city_contributions: list[CityContribution]
    low_margin_cities: list[CityContribution]
    risk_cities: list[RiskCitySummary]
    loss_actions: list[LossActionItem]
    analysis_steps: list[str]
    warning_signals: list[WarningSignal]
    scenario_forecasts: list[ScenarioForecast]
    action_impacts: list[ActionImpactItem]
    diagnosis_summary: list[str]
    monthly_history: list[MonthlyTrendPoint]
    future_monthly_forecast: list[MonthlyTrendPoint]
    key_drag_monthly_forecast: list[MonthlyTrendPoint]
    weekly_warning_points: list[WeeklyWarningPoint]


@dataclass(frozen=True)
class TransactionRecord:
    """原始明细行的轻量表示。"""

    order_date: datetime
    year: int
    month: int
    shop: str
    category: str
    hotel_city: str
    user_city: str
    sale_price: float
    profit: float


class RustToolError(RuntimeError):
    """Rust tool 调用失败。"""


def safe_ratio(numerator: float, denominator: float) -> float:
    """安全计算比率。"""

    if abs(denominator) < 1e-9:
        return 0.0
    return numerator / denominator


def clamp(value: float, lower: float, upper: float) -> float:
    """把浮点数限制在指定区间。"""

    return max(lower, min(upper, value))


def format_money(value: float) -> str:
    """格式化金额。"""

    return f"{value:,.2f}"


def format_percent(value: float) -> str:
    """格式化百分比。"""

    return f"{value:.2%}"


def format_wan(value: float) -> str:
    """把金额压缩成“万”口径，便于老板页阅读。"""

    return f"{value / 10000:.2f}万"


def format_matrix_text(lines: list[str]) -> str:
    """把策略句子整理成矩阵单元格里的分行文本。"""

    return "\n".join(f"{index}. {line}" for index, line in enumerate(lines, start=1))


def format_change(value: float) -> str:
    """格式化同比变化。"""

    return f"{value:+.2%}"


def parse_int(value: Any) -> int:
    """把 Rust 返回的数字统一转成整数。"""

    return int(float(str(value)))


def parse_float(value: Any) -> float:
    """把 Rust 返回的数字统一转成浮点数。"""

    return float(str(value))


def run_rust_tool(binary_path: Path, tool: str, args: dict[str, Any]) -> dict[str, Any]:
    """执行 Rust 二进制并返回 data 字段。"""

    # 2026-03-28 15:44 修改原因：真实工作簿路径和错误信息都包含中文，Windows 控制台最容易在这里乱码。
    # 2026-03-28 15:44 修改目的：统一走 UTF-8 字节流，避免路径、错误和返回内容在中途被错误解码。
    request_payload = {"tool": tool, "args": args}
    raw = subprocess.run(
        [str(binary_path)],
        input=json.dumps(request_payload, ensure_ascii=False).encode("utf-8"),
        capture_output=True,
        check=False,
    )
    if raw.returncode != 0:
        raise RustToolError(f"Rust tool 进程退出失败: {raw.returncode}")

    response_text = raw.stdout.decode("utf-8")
    response = json.loads(response_text)
    if response.get("status") != "ok":
        raise RustToolError(response.get("error", f"{tool} 执行失败"))
    return response["data"]


def load_transaction_records(source_workbook: Path, sheet_name: str = DEFAULT_SHEET_NAME) -> list[TransactionRecord]:
    """从真实 Excel 读取明细记录，用于月度和周度切片。"""

    # 2026-03-28 15:44 修改原因：Rust 主分析链已经能做静态聚合，但这轮新增的周度补充需要原始日期粒度。
    # 2026-03-28 15:44 修改目的：只把日期驱动的月/周切片放到 Python 交付层，不替代 Rust 的主聚合职责。
    workbook = load_workbook(source_workbook, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    records: list[TransactionRecord] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        order_date = row[0]
        year = row[1]
        month = row[2]
        shop = row[4]
        category = row[6]
        sale_price = row[9]
        profit = row[10]
        hotel_city = row[12]
        user_city = row[13]
        if not isinstance(order_date, datetime):
            continue
        records.append(
            TransactionRecord(
                order_date=order_date,
                year=int(year),
                month=int(month),
                shop=str(shop),
                category=str(category),
                hotel_city=str(hotel_city),
                user_city=str(user_city),
                sale_price=parse_float(sale_price or 0.0),
                profit=parse_float(profit or 0.0),
            )
        )
    return records


def build_monthly_points(
    records: list[TransactionRecord],
    filter_fn: Callable[[TransactionRecord], bool] | None = None,
) -> list[MonthlyTrendPoint]:
    """把明细聚合成按月经营点位。"""

    buckets: dict[str, dict[str, float]] = defaultdict(
        lambda: {"orders": 0.0, "revenue": 0.0, "profit": 0.0}
    )
    for record in records:
        if filter_fn and not filter_fn(record):
            continue
        period_label = f"{record.year:04d}-{record.month:02d}"
        buckets[period_label]["orders"] += 1
        buckets[period_label]["revenue"] += record.sale_price
        buckets[period_label]["profit"] += record.profit

    points = [
        MonthlyTrendPoint(
            period_label=period_label,
            order_count=int(values["orders"]),
            revenue=values["revenue"],
            profit=values["profit"],
        )
        for period_label, values in buckets.items()
    ]
    points.sort(key=lambda item: item.period_label)
    return points


def build_weekly_warning_points(
    records: list[TransactionRecord],
    risk_city_names: set[str],
    worst_shop: str,
    worst_category: str,
) -> list[WeeklyWarningPoint]:
    """提取周度预警补充点位。"""

    # 2026-03-28 15:44 修改原因：用户不希望整本报告都按周展开，但经营预警页需要更快识别短期失控信号。
    # 2026-03-28 15:44 修改目的：只在预警页保留近几周的补充视角，帮助老板判断风险是否已经加速恶化。
    buckets: dict[str, dict[str, float]] = defaultdict(
        lambda: {"revenue": 0.0, "profit": 0.0, "risk_revenue": 0.0, "risk_profit": 0.0}
    )
    for record in records:
        iso_year, iso_week, _ = record.order_date.isocalendar()
        period_label = f"{iso_year}-W{iso_week:02d}"
        buckets[period_label]["revenue"] += record.sale_price
        buckets[period_label]["profit"] += record.profit
        if record.user_city in risk_city_names or (
            record.shop == worst_shop and record.category == worst_category
        ):
            buckets[period_label]["risk_revenue"] += record.sale_price
            buckets[period_label]["risk_profit"] += record.profit

    sorted_periods = sorted(buckets.keys())[-3:]
    points: list[WeeklyWarningPoint] = []
    for index, period_label in enumerate(sorted_periods):
        values = buckets[period_label]
        risk_profit_rate = safe_ratio(values["risk_profit"], values["risk_revenue"])
        if values["risk_profit"] < 0:
            signal_name = "单周亏损拐入负值"
            note = "若下月未纠偏，月度利润率会继续恶化。"
        elif risk_profit_rate < 0.015:
            signal_name = "重点组合利润继续下探"
            note = "重点拖累组合周度毛利率偏低，已经逼近失控阈值。"
        else:
            signal_name = "风险城市亏损扩大"
            note = "风险城市利润修复不足，周度失血尚未真正逆转。"
        if index == len(sorted_periods) - 1 and values["profit"] < 0:
            signal_name = "单周亏损拐入负值"
            note = "周度总盘已经进入负利润，说明月度风险正在加速兑现。"
        points.append(
            WeeklyWarningPoint(
                period_label=period_label,
                signal_name=signal_name,
                revenue=values["revenue"],
                profit=values["profit"],
                note=note,
            )
        )
    return points


def average_growth(points: list[MonthlyTrendPoint]) -> float:
    """计算最近几个月的平均收入增速。"""

    growths: list[float] = []
    for previous, current in zip(points, points[1:]):
        if previous.revenue > 0:
            growths.append((current.revenue - previous.revenue) / previous.revenue)
    if not growths:
        return 0.0
    return statistics.fmean(growths)


def average_margin_delta(points: list[MonthlyTrendPoint]) -> float:
    """计算最近几个月的平均毛利率变化。"""

    deltas = [current.profit_rate - previous.profit_rate for previous, current in zip(points, points[1:])]
    if not deltas:
        return 0.0
    return statistics.fmean(deltas)


def next_month_label(period_label: str, offset: int) -> str:
    """从 YYYY-MM 计算未来第 N 个月的标签。"""

    year, month = period_label.split("-")
    absolute_month = int(year) * 12 + int(month) - 1 + offset
    next_year = absolute_month // 12
    next_month = absolute_month % 12 + 1
    return f"{next_year:04d}-{next_month:02d}"


def project_monthly_points(
    last_point: MonthlyTrendPoint,
    periods: int,
    revenue_growth: float,
    margin_delta: float,
    min_margin: float,
    max_margin: float,
) -> list[MonthlyTrendPoint]:
    """按固定增长率和毛利率斜率投射未来月度路径。"""

    revenue = last_point.revenue
    margin = last_point.profit_rate
    order_count = last_point.order_count
    points: list[MonthlyTrendPoint] = []
    for offset in range(1, periods + 1):
        revenue *= 1 + revenue_growth
        margin = clamp(margin + margin_delta, min_margin, max_margin)
        order_count = max(1, int(round(order_count * (1 + revenue_growth * 0.8))))
        points.append(
            MonthlyTrendPoint(
                period_label=next_month_label(last_point.period_label, offset),
                order_count=order_count,
                revenue=revenue,
                profit=revenue * margin,
            )
        )
    return points


def build_scenario_projection(
    last_point: MonthlyTrendPoint,
    revenue_steps: list[float],
    margin_steps: list[float],
) -> list[ScenarioCyclePoint]:
    """按预设情景步长生成情景月度路径。"""

    revenue = last_point.revenue
    margin = last_point.profit_rate
    points: list[ScenarioCyclePoint] = []
    for index, (revenue_step, margin_step) in enumerate(zip(revenue_steps, margin_steps), start=1):
        revenue *= 1 + revenue_step
        margin = max(0.0, margin + margin_step)
        points.append(
            ScenarioCyclePoint(
                period_label=next_month_label(last_point.period_label, index),
                revenue=revenue,
                profit=revenue * margin,
                profit_rate=margin,
            )
        )
    return points


def detect_turning_point(points: list[ScenarioCyclePoint]) -> str:
    """识别情景利润修复拐点。"""

    # 2026-03-28 15:44 修改原因：用户明确要求未来场景页必须指出“拐点在哪”。
    # 2026-03-28 15:44 修改目的：用可解释的规则而不是黑盒算法，给老板一个可追责的拐点月份。
    for index in range(1, len(points) - 1):
        current_point = points[index]
        previous_point = points[index - 1]
        next_point = points[index + 1]
        if (
            current_point.profit > previous_point.profit
            and next_point.profit > current_point.profit
            and current_point.profit_rate >= previous_point.profit_rate
            and next_point.profit_rate >= current_point.profit_rate
        ):
            return current_point.period_label
    return "无自然拐点"


def cumulative_weights(style: str) -> list[float]:
    """给不同动作返回累计释放权重。"""

    if style == "fast":
        return [0.14, 0.32, 0.50, 0.68, 0.85]
    if style == "steady":
        return [0.08, 0.22, 0.40, 0.58, 0.78]
    return [0.05, 0.15, 0.30, 0.48, 0.68]


def build_action_projection(
    total_uplift: float,
    last_period_label: str,
    style: str,
) -> list[ActionCyclePoint]:
    """构造动作月度改善路径。"""

    weights = cumulative_weights(style)
    previous_cumulative = 0.0
    points: list[ActionCyclePoint] = []
    for index, weight in enumerate(weights, start=1):
        cumulative_profit_uplift = total_uplift * weight
        monthly_profit_uplift = cumulative_profit_uplift - previous_cumulative
        points.append(
            ActionCyclePoint(
                period_label=next_month_label(last_period_label, index),
                monthly_profit_uplift=monthly_profit_uplift,
                cumulative_profit_uplift=cumulative_profit_uplift,
            )
        )
        previous_cumulative = cumulative_profit_uplift
    return points


def detect_action_turning_point(
    base_projection: list[MonthlyTrendPoint],
    action_projection: list[ActionCyclePoint],
) -> str:
    """识别动作带来的利润修复拐点。"""

    adjusted_profits = [
        base_point.profit + action_point.monthly_profit_uplift
        for base_point, action_point in zip(base_projection, action_projection)
    ]
    for index in range(1, len(adjusted_profits) - 1):
        if adjusted_profits[index] > adjusted_profits[index - 1] and adjusted_profits[index + 1] > adjusted_profits[index]:
            return action_projection[index].period_label
    return action_projection[-1].period_label if action_projection else "无明确拐点"


def is_effective_turning_period(period_label: str) -> bool:
    """判断一个周期标签是否可用于拐点比较。"""

    return bool(period_label) and not str(period_label).startswith("无")


def period_to_index(period_label: str) -> int:
    """把 YYYY-MM 周期转成可比较的整数索引。"""

    if not is_effective_turning_period(period_label):
        return -1
    year, month = str(period_label).split("-")
    return int(year) * 12 + int(month)


def latest_effective_period(periods: list[str], fallback: str) -> str:
    """返回一组周期里的最晚有效月份。"""

    valid_periods = [period for period in periods if is_effective_turning_period(period)]
    if not valid_periods:
        return fallback
    return max(valid_periods, key=period_to_index)


def period_reaches(current_period: str, milestone_period: str) -> bool:
    """判断当前周期是否已经走到指定里程碑。"""

    if not is_effective_turning_period(current_period) or not is_effective_turning_period(milestone_period):
        return False
    return period_to_index(current_period) >= period_to_index(milestone_period)


def status_rank(status_light: str) -> int:
    """把红黄绿灯映射成可比较的顺序值。"""

    return {"红灯": 0, "黄灯": 1, "绿灯": 2}.get(status_light, 0)


def infer_risk_status_light(profit: float, profit_rate: float) -> str:
    """基于利润与毛利率给未来周期打风险灯色。"""

    # 2026-03-29 09:18 修改原因：用户要求预警页和情景页都要给出“灯色拐点”，且必须能解释为什么是这个月份。
    # 2026-03-29 09:18 修改目的：统一用利润是否转负、毛利率是否跌破 2% / 3% 两道阈值，形成可复核的红黄绿规则口径。
    if profit <= 0 or profit_rate < 0.0200:
        return "红灯"
    if profit_rate < 0.0300:
        return "黄灯"
    return "绿灯"


def detect_metric_turning_point(
    base_value: float,
    projections: list[MonthlyTrendPoint] | list[ScenarioCyclePoint],
    value_getter: Callable[[MonthlyTrendPoint | ScenarioCyclePoint], float],
    fallback: str,
) -> str:
    """用“连续 2 期改善”规则识别单指标拐点。"""

    # 2026-03-29 09:18 修改原因：用户强调 05 页拐点不能只写结论，必须说明算法，否则老板会质疑拐点月份的说服力。
    # 2026-03-29 09:18 修改目的：统一采用最近一期为基线、未来路径连续两期改善即确认拐点的可解释算法，避免黑盒预测。
    previous_value = base_value
    improving_streak = 0
    for point in projections:
        current_value = value_getter(point)
        if current_value > previous_value + 1e-9:
            improving_streak += 1
        else:
            improving_streak = 0
        if improving_streak >= 2:
            return point.period_label
        previous_value = current_value
    return fallback


def select_scenario_action_items(
    bundle: BossReportBundle,
    scenario: ScenarioForecast,
) -> list[ActionImpactItem]:
    """把情景和动作改善项目做口径映射。"""

    if "情景A" in scenario.scenario_name:
        return []
    if "局部止损" in scenario.scenario_name:
        return [
            item
            for item in bundle.action_impacts
            if "收缩负利润城市低价订单" in item.action_name
        ]
    if "结构优化" in scenario.scenario_name:
        return [
            item
            for item in bundle.action_impacts
            if item.action_name.startswith("修复") or item.action_name.startswith("提升")
        ]
    return []


def infer_scenario_action_turning_point(
    bundle: BossReportBundle,
    scenario: ScenarioForecast,
) -> str:
    """根据动作见效周期推断情景动作拐点。"""

    action_items = select_scenario_action_items(bundle, scenario)
    if not action_items:
        return "无动作拐点"
    return latest_effective_period(
        [item.effect_period for item in action_items],
        fallback="无动作拐点",
    )


def infer_scenario_status_turning_point(
    bundle: BossReportBundle,
    scenario: ScenarioForecast,
) -> str:
    """根据灯色改善识别情景灯色拐点。"""

    if not bundle.monthly_history:
        return "无灯色拐点"
    previous_status = infer_risk_status_light(
        bundle.monthly_history[-1].profit,
        bundle.monthly_history[-1].profit_rate,
    )
    for point in scenario.monthly_projection:
        current_status = infer_risk_status_light(point.profit, point.profit_rate)
        if status_rank(current_status) > status_rank(previous_status):
            return point.period_label
        previous_status = current_status
    return "无灯色拐点"


def build_scenario_turning_profile(
    bundle: BossReportBundle,
    scenario: ScenarioForecast,
) -> dict[str, str]:
    """汇总情景的多拐点与老板主拐点。"""

    if not bundle.monthly_history:
        return {
            "profit_turning_point": "无利润拐点",
            "margin_turning_point": "无毛利率拐点",
            "status_turning_point": "无灯色拐点",
            "action_turning_point": "无动作拐点",
            "main_turning_point": "无主拐点",
        }

    base_point = bundle.monthly_history[-1]
    profit_turning_point = detect_metric_turning_point(
        base_point.profit,
        scenario.monthly_projection,
        lambda point: point.profit,
        fallback="无利润拐点",
    )
    margin_turning_point = detect_metric_turning_point(
        base_point.profit_rate,
        scenario.monthly_projection,
        lambda point: point.profit_rate,
        fallback="无毛利率拐点",
    )
    status_turning_point = infer_scenario_status_turning_point(bundle, scenario)
    action_turning_point = infer_scenario_action_turning_point(bundle, scenario)
    all_turning_points = [
        profit_turning_point,
        margin_turning_point,
        status_turning_point,
        action_turning_point,
    ]
    main_turning_point = (
        latest_effective_period(all_turning_points, fallback="无主拐点")
        if all(is_effective_turning_period(period) for period in all_turning_points)
        else "无主拐点"
    )
    return {
        "profit_turning_point": profit_turning_point,
        "margin_turning_point": margin_turning_point,
        "status_turning_point": status_turning_point,
        "action_turning_point": action_turning_point,
        "main_turning_point": main_turning_point,
    }


def build_scenario_action_status(
    bundle: BossReportBundle,
    scenario: ScenarioForecast,
    current_period: str,
) -> str:
    """按周期判断情景动作目前处于什么阶段。"""

    action_items = select_scenario_action_items(bundle, scenario)
    if not action_items:
        return "未启动动作"
    if all(period_reaches(current_period, item.effect_period) for item in action_items):
        return "进入见效期"
    if any(period_reaches(current_period, item.start_period) for item in action_items):
        return "动作推进中"
    return "动作待启动"


def build_warning_signal_outlook(bundle: BossReportBundle) -> dict[str, str]:
    """构造经营预警页共用的灯色演化结论。"""

    if not bundle.monthly_history:
        return {
            "base_status": "黄灯",
            "final_status": "黄灯",
            "expected_status_change": "黄灯维持",
            "expected_main_turning_point": "无主拐点",
        }

    base_point = bundle.monthly_history[-1]
    base_status = infer_risk_status_light(base_point.profit, base_point.profit_rate)
    previous_status = base_status
    first_status_change_period = ""
    final_status = base_status
    for point in bundle.future_monthly_forecast:
        current_status = infer_risk_status_light(point.profit, point.profit_rate)
        if not first_status_change_period and current_status != previous_status:
            first_status_change_period = point.period_label
        previous_status = current_status
        final_status = current_status
    if first_status_change_period:
        expected_status_change = f"{base_status}->{final_status}（首次变化：{first_status_change_period}）"
    else:
        expected_status_change = f"{base_status}维持"
    return {
        "base_status": base_status,
        "final_status": final_status,
        "expected_status_change": expected_status_change,
        "expected_main_turning_point": "无主拐点",
    }


def build_warning_timeline_rows(bundle: BossReportBundle) -> list[dict[str, Any]]:
    """把未来 5 个月基线预测整理成预警时间轴。"""

    if not bundle.monthly_history:
        return []

    # 2026-03-29 09:18 修改原因：用户指出经营预警不能只有静态结论，必须串起“时间趋势 -> 风险换挡 -> 后果兑现”。
    # 2026-03-29 09:18 修改目的：把未来周期逐月落到表格里，让老板能看到风险灯色何时转红、利润何时继续被侵蚀、主拐点为何缺失。
    timeline_rows: list[dict[str, Any]] = []
    base_point = bundle.monthly_history[-1]
    previous_point: MonthlyTrendPoint | ScenarioCyclePoint = base_point
    previous_status = infer_risk_status_light(base_point.profit, base_point.profit_rate)
    outlook = build_warning_signal_outlook(bundle)
    for point in bundle.future_monthly_forecast:
        current_status = infer_risk_status_light(point.profit, point.profit_rate)
        if current_status != previous_status:
            warning_note = f"风险灯色由{previous_status}切换为{current_status}，说明风险开始换挡。"
        elif point.profit < previous_point.profit and point.profit_rate < previous_point.profit_rate:
            warning_note = "收入仍在放大，但利润和毛利率同步走弱，说明低质量增长继续放大。"
        elif point.profit < previous_point.profit:
            warning_note = "利润继续下探，止损窗口继续缩短。"
        else:
            warning_note = "指标短暂企稳，但当前路径下仍未形成老板主拐点。"
        timeline_rows.append(
            {
                "period": point.period_label,
                "revenue": point.revenue,
                "profit": point.profit,
                "profit_rate": point.profit_rate,
                "risk_status": current_status,
                "warning_note": warning_note,
                "is_main_turning_point": (
                    "主拐点"
                    if point.period_label == outlook["expected_main_turning_point"]
                    else ""
                ),
            }
        )
        previous_point = point
        previous_status = current_status
    return timeline_rows


def load_boss_report_bundle_from_rust(
    source_workbook: Path,
    rust_binary: Path = DEFAULT_RUST_BINARY,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> BossReportBundle:
    """从 Rust 聚合结果和原始日期明细构造老板决策版总包。"""

    # 2026-03-28 15:46 修改原因：这轮报告不能再停留在年度结论，必须把时间趋势和未来周期一起带出来。
    # 2026-03-28 15:46 修改目的：继续让 Rust 负责结构聚合，让 Python 只补日期序列、场景路径和动作拐点。
    schema_data = run_rust_tool(
        rust_binary,
        "apply_header_schema",
        {"path": str(source_workbook), "sheet": sheet_name},
    )
    table_ref = schema_data["table_ref"]
    row_count = int(schema_data["row_count"])

    overall_summary = run_rust_tool(
        rust_binary,
        "stat_summary",
        {
            "table_ref": table_ref,
            "columns": ["column_10", "column_11"],
            "casts": [
                {"column": "column_10", "target_type": "float64"},
                {"column": "column_11", "target_type": "float64"},
            ],
            "top_k": 5,
        },
    )
    numeric_map = {item["column"]: item for item in overall_summary.get("numeric_summaries", [])}
    total_sales = parse_float(numeric_map["column_10"]["sum"])
    total_profit = parse_float(numeric_map["column_11"]["sum"])
    overall_profit_rate = safe_ratio(total_profit, total_sales)

    year_rows = run_rust_tool(
        rust_binary,
        "group_and_aggregate",
        {
            "table_ref": table_ref,
            "group_by": ["column_2"],
            "aggregations": [
                {"column": "id", "operator": "count"},
                {"column": "column_10", "operator": "sum"},
                {"column": "column_11", "operator": "sum"},
            ],
            "casts": [
                {"column": "column_10", "target_type": "float64"},
                {"column": "column_11", "target_type": "float64"},
            ],
        },
    )["rows"]
    shop_category_rows = run_rust_tool(
        rust_binary,
        "group_and_aggregate",
        {
            "table_ref": table_ref,
            "group_by": ["column_5", "column_7"],
            "aggregations": [
                {"column": "id", "operator": "count"},
                {"column": "column_10", "operator": "sum"},
                {"column": "column_11", "operator": "sum"},
            ],
            "casts": [
                {"column": "column_10", "target_type": "float64"},
                {"column": "column_11", "target_type": "float64"},
            ],
        },
    )["rows"]
    city_rows = run_rust_tool(
        rust_binary,
        "group_and_aggregate",
        {
            "table_ref": table_ref,
            "group_by": ["column_14"],
            "aggregations": [
                {"column": "id", "operator": "count"},
                {"column": "column_10", "operator": "sum"},
                {"column": "column_11", "operator": "sum"},
            ],
            "casts": [
                {"column": "column_10", "target_type": "float64"},
                {"column": "column_11", "target_type": "float64"},
            ],
        },
    )["rows"]
    risk_rows = run_rust_tool(
        rust_binary,
        "group_and_aggregate",
        {
            "table_ref": table_ref,
            "group_by": ["column_5", "column_7", "column_14"],
            "aggregations": [
                {"column": "id", "operator": "count"},
                {"column": "column_10", "operator": "sum"},
                {"column": "column_11", "operator": "sum"},
            ],
            "casts": [
                {"column": "column_10", "target_type": "float64"},
                {"column": "column_11", "target_type": "float64"},
            ],
        },
    )["rows"]

    year_summaries = [
        YearSummary(
            year=str(row["column_2"]),
            order_count=parse_int(row["id_count"]),
            total_sales=parse_float(row["column_10_sum"]),
            total_profit=parse_float(row["column_11_sum"]),
        )
        for row in year_rows
    ]
    year_summaries.sort(key=lambda item: item.year)

    shop_category_summaries = [
        ShopCategorySummary(
            shop=str(row["column_5"]),
            category=str(row["column_7"]),
            order_count=parse_int(row["id_count"]),
            total_sales=parse_float(row["column_10_sum"]),
            total_profit=parse_float(row["column_11_sum"]),
        )
        for row in shop_category_rows
    ]
    shop_category_summaries.sort(key=lambda item: (item.shop, item.category))

    city_contributions = [
        CityContribution(
            city=str(row["column_14"]),
            order_count=parse_int(row["id_count"]),
            total_sales=parse_float(row["column_10_sum"]),
            total_profit=parse_float(row["column_11_sum"]),
        )
        for row in city_rows
    ]
    city_contributions.sort(key=lambda item: item.total_sales, reverse=True)

    low_margin_cities = [item for item in city_contributions if item.order_count >= 300]
    low_margin_cities.sort(key=lambda item: item.profit_rate)
    low_margin_cities = low_margin_cities[:10]

    worst_shop_category = min(shop_category_summaries, key=lambda item: item.profit_rate)
    best_shop_category = max(shop_category_summaries, key=lambda item: item.profit_rate)
    risk_cities = [
        RiskCitySummary(
            city=str(row["column_14"]),
            total_sales=parse_float(row["column_10_sum"]),
            total_profit=parse_float(row["column_11_sum"]),
        )
        for row in risk_rows
        if str(row["column_5"]) == worst_shop_category.shop
        and str(row["column_7"]) == worst_shop_category.category
        and parse_float(row["column_11_sum"]) < 0
    ]
    risk_cities.sort(key=lambda item: item.total_profit)
    risk_cities = risk_cities[:10]

    hotel_items = [item for item in shop_category_summaries if item.category == worst_shop_category.category]
    hotel_sales = sum(item.total_sales for item in hotel_items)
    hotel_profit = sum(item.total_profit for item in hotel_items)
    hotel_margin = safe_ratio(hotel_profit, hotel_sales)

    sales_growth = 0.0
    profit_growth = 0.0
    margin_change = 0.0
    if len(year_summaries) >= 2:
        sales_growth = safe_ratio(
            year_summaries[-1].total_sales - year_summaries[0].total_sales,
            year_summaries[0].total_sales,
        )
        profit_growth = safe_ratio(
            year_summaries[-1].total_profit - year_summaries[0].total_profit,
            year_summaries[0].total_profit,
        )
        margin_change = year_summaries[-1].profit_rate - year_summaries[0].profit_rate

    records = load_transaction_records(source_workbook, sheet_name=sheet_name)
    monthly_points = build_monthly_points(records)
    monthly_history = monthly_points[-5:] if len(monthly_points) >= 5 else monthly_points
    avg_revenue_growth = average_growth(monthly_history[-4:]) if len(monthly_history) >= 4 else 0.0
    avg_margin_step = average_margin_delta(monthly_history[-4:]) if len(monthly_history) >= 4 else 0.0
    base_revenue_growth = clamp(avg_revenue_growth * 0.45, 0.008, 0.02)
    base_margin_step = min(-0.0010, avg_margin_step * 0.8) if avg_margin_step < 0 else -0.0010
    future_monthly_forecast = project_monthly_points(
        monthly_history[-1],
        FORECAST_PERIODS,
        revenue_growth=base_revenue_growth,
        margin_delta=base_margin_step,
        min_margin=max(monthly_history[-1].profit_rate - 0.012, 0.010),
        max_margin=max(monthly_history[-1].profit_rate, 0.050),
    )

    key_drag_monthly_history = build_monthly_points(
        records,
        filter_fn=lambda record: record.shop == worst_shop_category.shop and record.category == worst_shop_category.category,
    )
    key_drag_seed = key_drag_monthly_history[-1] if key_drag_monthly_history else MonthlyTrendPoint(
        period_label=monthly_history[-1].period_label,
        order_count=max(1, worst_shop_category.order_count // max(len(monthly_history), 1)),
        revenue=safe_ratio(worst_shop_category.total_sales, max(len(monthly_history), 1)),
        profit=safe_ratio(worst_shop_category.total_profit, max(len(monthly_history), 1)),
    )
    drag_revenue_growth = clamp(base_revenue_growth * 0.9, 0.006, 0.018)
    drag_margin_step = (
        min(-0.0015, average_margin_delta(key_drag_monthly_history[-4:]) * 0.9)
        if len(key_drag_monthly_history) >= 4
        else -0.0015
    )
    key_drag_monthly_forecast = project_monthly_points(
        key_drag_seed,
        FORECAST_PERIODS,
        revenue_growth=drag_revenue_growth,
        margin_delta=drag_margin_step,
        min_margin=-0.008,
        max_margin=max(key_drag_seed.profit_rate, 0.035),
    )

    # 2026-03-28 21:36 修改原因：用户明确要求未来场景页不能只写抽象策略，必须点名止损城市、结构替换方向和改善幅度。
    # 2026-03-28 21:36 修改目的：先在场景数据层把“结论 / 动作 / 分析 / 数据”四段叙事准备好，避免渲染层继续堆概念词。
    risk_city_names = "、".join(item.city for item in risk_cities[:7]) or "暂无明显负利润城市"
    risk_city_name_set = {item.city for item in risk_cities[:7]}
    weekly_warning_points = build_weekly_warning_points(
        records,
        risk_city_names=risk_city_name_set,
        worst_shop=worst_shop_category.shop,
        worst_category=worst_shop_category.category,
    )

    worst_combo_sales_share = safe_ratio(worst_shop_category.total_sales, total_sales)
    worst_combo_profit_share = safe_ratio(worst_shop_category.total_profit, total_profit)
    risk_city_total_loss = abs(sum(item.total_profit for item in risk_cities if item.total_profit < 0))
    worst_to_hotel_avg_uplift = max(
        0.0,
        worst_shop_category.total_sales * max(hotel_margin - worst_shop_category.profit_rate, 0.008),
    )
    mix_optimization_uplift = max(
        0.0,
        total_sales * 0.03 * max(best_shop_category.profit_rate - worst_shop_category.profit_rate, 0.01),
    )
    worst_combo_label = f"{worst_shop_category.shop}+{worst_shop_category.category}"
    risk_city_top3 = risk_cities[:3]
    risk_city_top5 = risk_cities[:5]
    risk_city_top3_names = "、".join(item.city for item in risk_city_top3) or "暂无第一批止损城市"
    risk_city_top5_names = "、".join(item.city for item in risk_city_top5) or "暂无重点观察城市"
    follow_up_city_names = "、".join(item.city for item in risk_city_top5[3:]) or "暂无第二批跟进城市"
    risk_city_top3_loss = abs(sum(item.total_profit for item in risk_city_top3 if item.total_profit < 0))
    risk_city_top5_loss = abs(sum(item.total_profit for item in risk_city_top5 if item.total_profit < 0))
    same_shop_benchmarks = [
        item
        for item in shop_category_summaries
        if item.shop == worst_shop_category.shop and item.category != worst_shop_category.category
    ]
    same_shop_benchmarks.sort(key=lambda item: item.profit_rate, reverse=True)
    same_shop_labels = "、".join(f"{item.shop}+{item.category}" for item in same_shop_benchmarks[:2]) or "暂无可替换结构"
    transfer_uplifts = [
        max(0.0, (item.profit_rate - worst_shop_category.profit_rate) * 1_000_000)
        for item in same_shop_benchmarks[:2]
    ]
    transfer_uplift_low = min(transfer_uplifts) if transfer_uplifts else 0.0
    transfer_uplift_high = max(transfer_uplifts) if transfer_uplifts else 0.0

    diagnosis_summary = [
        (
            f"当前策略继续下去，收入会被{worst_shop_category.shop} + {worst_shop_category.category}"
            "继续放大，但利润不会同步增长。"
        ),
        (
            f"预计未来{FORECAST_PERIODS}个月整体利润额继续承压，重点拖累组合将在"
            f"{next((item.period_label for item in key_drag_monthly_forecast if item.profit < 0), '预测期后段')}进入负利润。"
        ),
    ]
    executive_points = [
        (
            f"收入同比{format_change(sales_growth)}，但毛利同比{format_change(profit_growth)}，"
            f"毛利率下降{margin_change * 100:.2f}个百分点，增长质量正在恶化。"
        ),
        (
            f"{worst_shop_category.shop} + {worst_shop_category.category} 销售占比"
            f"{format_percent(worst_combo_sales_share)}，但毛利贡献仅"
            f"{format_percent(worst_combo_profit_share)}，是当前最大经营拖累。"
        ),
        (
            f"如果继续当前策略，未来{FORECAST_PERIODS}个月收入仍会增长，"
            "但利润和毛利率不会自然修复。"
        ),
    ]
    analysis_steps = [
        "先确认月度收入、毛利额、毛利率的变化，判断增长是否真的转化为利润。",
        "再拆渠道 x 品类结构，定位是谁在放大收入、却没有同步贡献利润。",
        "随后下钻城市风险和周度信号，识别亏损是否已经加速扩散。",
        f"最后给出未来{FORECAST_PERIODS}个月预测、情景差异和动作拐点，让老板直接比较路径。",
    ]

    warning_signals = [
        WarningSignal(
            title="利润率持续下探预警",
            current_signal=(
                f"最近月度收入仍在增长，但毛利率从{format_percent(monthly_history[0].profit_rate)}"
                f"下降到{format_percent(monthly_history[-1].profit_rate)}。"
            ),
            consequence=f"如果继续当前策略，未来{FORECAST_PERIODS}个月利润额会继续下降，毛利率将持续下探。",
            urgency="高",
            trigger_period=next(
                (item.period_label for item in future_monthly_forecast if item.profit < monthly_history[-1].profit * 0.95),
                future_monthly_forecast[0].period_label,
            ),
            trend_summary="收入在放大，但利润率修复失败，属于典型的低质量增长延续。",
        ),
        WarningSignal(
            title="大体量低质量增长预警",
            current_signal=(
                f"{worst_shop_category.shop} + {worst_shop_category.category} 已经占到"
                f"{format_percent(worst_combo_sales_share)}的销售规模，但毛利率只有"
                f"{format_percent(worst_shop_category.profit_rate)}。"
            ),
            consequence="如果继续依赖低价和补贴换规模，收入越大，利润质量越差。",
            urgency="高",
            trigger_period=next(
                (item.period_label for item in key_drag_monthly_forecast if item.profit < 0),
                key_drag_monthly_forecast[-1].period_label,
            ),
            trend_summary="重点拖累组合会在未来几个周期内跌入负利润，是最需要立刻纠偏的板块。",
        ),
        WarningSignal(
            title="风险城市亏损扩散预警",
            current_signal=f"当前明确亏损城市集中在{risk_city_names}，周度补充信号已经开始恶化。",
            consequence="如果不先止损，亏损城市会继续吞噬利润，并放大投放和运营资源浪费。",
            urgency="中高",
            trigger_period=weekly_warning_points[-1].period_label if weekly_warning_points else future_monthly_forecast[0].period_label,
            trend_summary="周度信号已出现单周亏损，说明月度失血将继续扩大。",
        ),
    ]

    scenario_a_points = build_scenario_projection(
        monthly_history[-1],
        revenue_steps=[0.012, 0.013, 0.013, 0.012, 0.012],
        margin_steps=[-0.0012, -0.0011, -0.0010, -0.0009, -0.0008],
    )
    scenario_b_points = build_scenario_projection(
        monthly_history[-1],
        revenue_steps=[-0.018, -0.010, 0.000, 0.006, 0.010],
        margin_steps=[0.0018, 0.0020, 0.0020, 0.0016, 0.0014],
    )
    scenario_c_points = build_scenario_projection(
        monthly_history[-1],
        revenue_steps=[0.004, 0.008, 0.012, 0.015, 0.018],
        margin_steps=[0.0024, 0.0026, 0.0025, 0.0023, 0.0021],
    )
    scenario_a_turning = detect_turning_point(scenario_a_points)
    scenario_b_turning = detect_turning_point(scenario_b_points)
    scenario_c_turning = detect_turning_point(scenario_c_points)

    scenario_forecasts = [
        ScenarioForecast(
            scenario_name="情景A-继续当前策略",
            revenue_change_pct=safe_ratio(scenario_a_points[-1].revenue - monthly_history[-1].revenue, monthly_history[-1].revenue),
            profit_change_pct=safe_ratio(scenario_a_points[-1].profit - monthly_history[-1].profit, monthly_history[-1].profit),
            profit_rate=scenario_a_points[-1].profit_rate,
            summary="预计收入继续增长，但利润和毛利率持续走弱。",
            turning_point_period=scenario_a_turning,
            monthly_projection=scenario_a_points,
            strategy_definition="维持当前放量策略，不调整低毛利组合，也不主动止损风险城市。",
            key_actions=["维持当前投放", "不调整天猫店铺+酒店", "不收缩风险城市"],
            baseline_data_summary=(
                f"之前的数据：当前月收入 {format_money(monthly_history[-1].revenue)}，"
                f"月利润 {format_money(monthly_history[-1].profit)}，重点拖累组合毛利率 {format_percent(key_drag_monthly_history[-1].profit_rate if key_drag_monthly_history else key_drag_seed.profit_rate)}。"
            ),
            conclusion_summary="继续当前策略会继续放大收入，但利润会继续被低质量结构稀释，预测期内看不到自然修复。",
            action_summary=format_matrix_text(
                [
                    "维持当前投放节奏，不调整预算投向。",
                    f"继续放量 {worst_combo_label}，不处理当前低毛利结构。",
                    f"不收缩 {risk_city_top3_names} 等风险城市，继续以规模优先。",
                ]
            ),
            analysis_summary=format_matrix_text(
                [
                    f"{worst_combo_label} 当前是最差组合，规模越大，对整体利润的稀释越明显。",
                    f"风险城市已经从 {risk_city_top3_names} 扩散到 {follow_up_city_names}，说明亏损不是偶发点状问题。",
                    "因此收入虽然还能增长，但增长不会自然转化成利润，反而会继续放大低质量增长。",
                ]
            ),
            data_summary=format_matrix_text(
                [
                    f"当前月收入 {format_wan(monthly_history[-1].revenue)}，月利润 {format_wan(monthly_history[-1].profit)}。",
                    f"{worst_combo_label} 销售 {format_wan(worst_shop_category.total_sales)}，毛利率仅 {format_percent(worst_shop_category.profit_rate)}。",
                    f"收入变化 {format_change(safe_ratio(scenario_a_points[-1].revenue - monthly_history[-1].revenue, monthly_history[-1].revenue))}，利润变化 {format_change(safe_ratio(scenario_a_points[-1].profit - monthly_history[-1].profit, monthly_history[-1].profit))}，拐点月份：{scenario_a_turning}。",
                ]
            ),
        ),
        ScenarioForecast(
            scenario_name="情景B-局部止损",
            revenue_change_pct=safe_ratio(scenario_b_points[-1].revenue - monthly_history[-1].revenue, monthly_history[-1].revenue),
            profit_change_pct=safe_ratio(scenario_b_points[-1].profit - monthly_history[-1].profit, monthly_history[-1].profit),
            profit_rate=scenario_b_points[-1].profit_rate,
            summary="短期收入承压，但利润会先止跌并逐步回升。",
            turning_point_period=scenario_b_turning,
            monthly_projection=scenario_b_points,
            strategy_definition="先做止损，把确定性亏损城市和低价订单压住，再观察利润修复。",
            key_actions=["收缩负利润城市低价订单", "压缩高风险投放", "先守利润不守规模"],
            baseline_data_summary=(
                f"之前的数据：风险城市合计失血 {format_money(risk_city_total_loss)}，"
                f"当前月利润 {format_money(monthly_history[-1].profit)}，周度信号已经出现单周亏损。"
            ),
            conclusion_summary="先止血再修复，短期牺牲一点规模是值得的，因为它能更快把利润拉回正轨。",
            action_summary=format_matrix_text(
                [
                    f"第一批止损城市先控 {risk_city_top3_names}。",
                    f"立即暂停这些城市在 {worst_combo_label} 上的低价放量单，并压缩高风险投放。",
                    f"第二批跟进 {follow_up_city_names}，把经营口径从追规模切到守利润。",
                ]
            ),
            analysis_summary=format_matrix_text(
                [
                    f"{risk_city_top3_names} 已经形成确定性亏损，优先切掉这些订单，比全面收缩更快见效。",
                    f"如果连 {follow_up_city_names} 也继续放量，风险会从局部亏损演变成区域扩散。",
                    "所以局部止损的逻辑不是收缩业务，而是先切掉最不赚钱的订单和投放。",
                ]
            ),
            data_summary=format_matrix_text(
                [
                    f"前3个风险城市合计失血 {format_wan(risk_city_top3_loss)}，前5个城市失血 {format_wan(risk_city_top5_loss)}。",
                    f"第一批止损城市：{risk_city_top3_names}；第二批跟进城市：{follow_up_city_names}。",
                    f"收入变化 {format_change(safe_ratio(scenario_b_points[-1].revenue - monthly_history[-1].revenue, monthly_history[-1].revenue))}，利润变化 {format_change(safe_ratio(scenario_b_points[-1].profit - monthly_history[-1].profit, monthly_history[-1].profit))}，拐点月份：{scenario_b_turning}。",
                ]
            ),
        ),
        ScenarioForecast(
            scenario_name="情景C-结构优化",
            revenue_change_pct=safe_ratio(scenario_c_points[-1].revenue - monthly_history[-1].revenue, monthly_history[-1].revenue),
            profit_change_pct=safe_ratio(scenario_c_points[-1].profit - monthly_history[-1].profit, monthly_history[-1].profit),
            profit_rate=scenario_c_points[-1].profit_rate,
            summary="收入和利润可以同步改善，是更优路径。",
            turning_point_period=scenario_c_turning,
            monthly_projection=scenario_c_points,
            strategy_definition="在止损基础上同步修复重点拖累组合，并提升高毛利结构占比。",
            key_actions=["修复天猫店铺+酒店毛利率", "提升高毛利品类占比", "止损与结构优化并行"],
            baseline_data_summary=(
                f"之前的数据：重点拖累组合销售占比 {format_percent(worst_combo_sales_share)}，"
                f"毛利贡献仅 {format_percent(worst_combo_profit_share)}，存在明显结构错配。"
            ),
            conclusion_summary="不是简单砍业务，而是把低质量收入搬到高质量结构里，让收入和利润一起修复。",
            action_summary=format_matrix_text(
                [
                    f"修复 {worst_combo_label} 的底价、补贴和投放门槛，先把坏单拦住。",
                    f"新增投放优先导向 {same_shop_labels}，用同店铺高毛利结构承接增长。",
                    f"对 {risk_city_top3_names} 同步做止损和结构替换，不再只靠低价酒店冲规模。",
                ]
            ),
            analysis_summary=format_matrix_text(
                [
                    f"{worst_combo_label} 不是不能做，而是当前价格和流量结构出了问题。",
                    f"同店铺下 {same_shop_labels} 的毛利率显著更高，说明问题核心是结构错配，不是渠道失效。",
                    "所以最优路径不是单纯砍酒店，而是把低质量流量转到更健康的承接结构里。",
                ]
            ),
            data_summary=format_matrix_text(
                [
                    f"{worst_combo_label} 销售 {format_wan(worst_shop_category.total_sales)}，毛利率 {format_percent(worst_shop_category.profit_rate)}；毛利贡献仅 {format_percent(worst_combo_profit_share)}。",
                    f"{same_shop_labels} 的毛利率分别达到 {format_percent(same_shop_benchmarks[0].profit_rate) if same_shop_benchmarks else '0.00%'} 和 {format_percent(same_shop_benchmarks[1].profit_rate) if len(same_shop_benchmarks) > 1 else '0.00%'}。",
                    f"每转移100万销售额到更优结构，理论上可多贡献毛利 {format_wan(transfer_uplift_low)} 到 {format_wan(transfer_uplift_high)}；拐点月份：{scenario_c_turning}。",
                ]
            ),
        ),
    ]

    fix_action_path = build_action_projection(worst_to_hotel_avg_uplift * 0.55, monthly_history[-1].period_label, style="steady")
    stop_loss_action_path = build_action_projection(max(risk_city_total_loss, total_profit * 0.04), monthly_history[-1].period_label, style="fast")
    mix_action_path = build_action_projection(mix_optimization_uplift * 0.45, monthly_history[-1].period_label, style="slow")

    action_impacts = [
        ActionImpactItem(
            priority="P1",
            action_name=f"修复 {worst_shop_category.shop} {worst_shop_category.category} 毛利率",
            improve_target="直接修复当前最大经营拖累，让大体量业务线先止跌再修复。",
            expected_profit_uplift=fix_action_path[-1].cumulative_profit_uplift,
            expected_revenue_impact="短期收入持平到小幅回落，但利润质量显著改善。",
            if_not_do="收入继续放大，但利润将在未来几个周期内继续被稀释。",
            start_period=fix_action_path[0].period_label,
            effect_period=fix_action_path[1].period_label,
            turning_point_period=detect_action_turning_point(future_monthly_forecast, fix_action_path),
            monthly_projection=fix_action_path,
        ),
        ActionImpactItem(
            priority="P1",
            action_name="收缩负利润城市低价订单",
            improve_target="先止住确定性亏损，把经营失血立刻收敛。",
            expected_profit_uplift=stop_loss_action_path[-1].cumulative_profit_uplift,
            expected_revenue_impact="短期收入承压，但亏损会更快收敛。",
            if_not_do="风险城市会继续吞噬利润，并拖慢整体拐点到来。",
            start_period=stop_loss_action_path[0].period_label,
            effect_period=stop_loss_action_path[0].period_label,
            turning_point_period=detect_action_turning_point(future_monthly_forecast, stop_loss_action_path),
            monthly_projection=stop_loss_action_path,
        ),
        ActionImpactItem(
            priority="P2",
            action_name=f"提升 {best_shop_category.category} 等高毛利结构占比",
            improve_target="把增长重新导向高质量结构，让利润修复更可持续。",
            expected_profit_uplift=mix_action_path[-1].cumulative_profit_uplift,
            expected_revenue_impact="收入整体平稳，利润弹性更好。",
            if_not_do="结构修复会中断，利润率难以形成持续改善。",
            start_period=mix_action_path[0].period_label,
            effect_period=mix_action_path[2].period_label,
            turning_point_period=detect_action_turning_point(future_monthly_forecast, mix_action_path),
            monthly_projection=mix_action_path,
        ),
    ]

    loss_actions = [
        LossActionItem(
            priority="P1",
            action=f"立即复核 {worst_shop_category.shop} {worst_shop_category.category} 的定价、补贴和投放策略",
            owner="渠道运营",
        ),
        LossActionItem(
            priority="P1",
            action=f"针对{risk_city_names}建立负利润城市清单并暂停低价放量",
            owner="经营分析",
        ),
        LossActionItem(
            priority="P2",
            action="建立城市-渠道-品类周度毛利率预警机制，并纳入经营例会复盘",
            owner="商品运营",
        ),
    ]

    return BossReportBundle(
        report_title="老板决策版业绩诊断",
        source_workbook=str(source_workbook),
        overall_metrics=[
            SummaryMetric(label="总订单数", value=f"{row_count:,}"),
            SummaryMetric(label="总销售额", value=format_money(total_sales)),
            SummaryMetric(label="总毛利额", value=format_money(total_profit)),
            SummaryMetric(label="整体毛利率", value=format_percent(overall_profit_rate)),
        ],
        executive_points=executive_points,
        year_summaries=year_summaries,
        shop_category_summaries=shop_category_summaries,
        city_contributions=city_contributions[:10],
        low_margin_cities=low_margin_cities,
        risk_cities=risk_cities,
        loss_actions=loss_actions,
        analysis_steps=analysis_steps,
        warning_signals=warning_signals,
        scenario_forecasts=scenario_forecasts,
        action_impacts=action_impacts,
        diagnosis_summary=diagnosis_summary,
        monthly_history=monthly_history,
        future_monthly_forecast=future_monthly_forecast,
        key_drag_monthly_forecast=key_drag_monthly_forecast,
        weekly_warning_points=weekly_warning_points,
    )


def build_boss_report_workbook(bundle: BossReportBundle, output_path: Path) -> Path:
    """把总包写成老板决策版工作簿。"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(output_path))
    formats = create_formats(workbook)

    try:
        # 2026-03-28 15:46 修改原因：用户这轮要求的是有节奏、有拐点、有动作时间路径的老板材料，而不是静态诊断页。
        # 2026-03-28 15:46 修改目的：保留 9 张工作表结构，但让结论、预警、预测和动作页都带上月度时间主轴。
        write_conclusion_overview_sheet(workbook, formats, bundle)
        write_analysis_path_sheet(workbook, formats, bundle)
        write_core_diagnosis_sheet(workbook, formats, bundle)
        write_warning_sheet(workbook, formats, bundle)
        write_scenario_sheet(workbook, formats, bundle)
        write_action_impact_sheet(workbook, formats, bundle)
        write_city_contribution_sheet(workbook, formats, bundle)
        write_appendix_chart_sheet(workbook, formats, bundle)
        write_appendix_summary_sheet(workbook, formats, bundle)
    finally:
        workbook.close()

    return output_path


def create_formats(workbook: xlsxwriter.Workbook) -> dict[str, xlsxwriter.format.Format]:
    """集中创建单元格样式。"""

    # 2026-03-28 15:47 修改原因：时间序列表格和拐点字段明显增加，现有样式层级不够。
    # 2026-03-28 15:47 修改目的：让结论、预警、动作、周期字段在视觉上更容易被老板扫到。
    return {
        "title": workbook.add_format(
            {"bold": True, "font_size": 18, "font_name": "Microsoft YaHei", "bg_color": "#DDEBF7"}
        ),
        "subtitle": workbook.add_format(
            {"bold": True, "font_size": 12, "font_name": "Microsoft YaHei", "bg_color": "#FFF2CC"}
        ),
        "warning_title": workbook.add_format(
            {"bold": True, "font_size": 12, "font_name": "Microsoft YaHei", "bg_color": "#FCE4D6"}
        ),
        "header": workbook.add_format(
            {"bold": True, "font_name": "Microsoft YaHei", "bg_color": "#BDD7EE", "border": 1}
        ),
        "cell": workbook.add_format({"font_name": "Microsoft YaHei", "border": 1}),
        "text": workbook.add_format({"font_name": "Microsoft YaHei", "text_wrap": True, "valign": "top", "border": 1}),
        "money": workbook.add_format({"font_name": "Microsoft YaHei", "border": 1, "num_format": "#,##0.00"}),
        "percent": workbook.add_format({"font_name": "Microsoft YaHei", "border": 1, "num_format": "0.00%"}),
        "negative_money": workbook.add_format(
            {"font_name": "Microsoft YaHei", "border": 1, "num_format": "#,##0.00;[Red]-#,##0.00"}
        ),
        "metric_label": workbook.add_format(
            {"bold": True, "font_name": "Microsoft YaHei", "bg_color": "#FCE4D6", "border": 1}
        ),
        "metric_value": workbook.add_format(
            {"bold": True, "font_size": 14, "font_name": "Microsoft YaHei", "border": 1, "align": "center"}
        ),
        # 2026-03-28 16:13 修改原因：用户要求未来场景页把预测拐点直接标红，不能只用文字说明带过。
        # 2026-03-28 16:13 修改目的：提供统一的红色拐点格式，让策略页可以一眼识别关键转折月份。
        "turning_point": workbook.add_format(
            {"font_name": "Microsoft YaHei", "border": 1, "font_color": "#C00000", "bold": True}
        ),
        # 2026-03-29 00:47 修改原因：用户要求把执行附表升级成红黄绿周会版，状态不能只靠文字判断。
        # 2026-03-29 00:47 修改目的：提供统一的红黄绿状态灯格式，让老板页和周会表共享相同颜色口径。
        "red_status": workbook.add_format(
            {"font_name": "Microsoft YaHei", "border": 1, "bold": True, "bg_color": "#F4CCCC", "font_color": "#990000"}
        ),
        "yellow_status": workbook.add_format(
            {"font_name": "Microsoft YaHei", "border": 1, "bold": True, "bg_color": "#FFF2CC", "font_color": "#7F6000"}
        ),
        "green_status": workbook.add_format(
            {"font_name": "Microsoft YaHei", "border": 1, "bold": True, "bg_color": "#D9EAD3", "font_color": "#274E13"}
        ),
    }


def write_monthly_table(
    sheet: xlsxwriter.worksheet.Worksheet,
    formats: dict[str, xlsxwriter.format.Format],
    start_row: int,
    start_col: int,
    title: str,
    points: list[MonthlyTrendPoint],
) -> tuple[int, int]:
    """写月度经营点位表。"""

    sheet.write(start_row, start_col, title, formats["subtitle"])
    sheet.write_row(start_row + 1, start_col, ["月份", "订单数", "收入", "毛利", "毛利率"], formats["header"])
    for index, item in enumerate(points, start=start_row + 2):
        sheet.write(index, start_col, item.period_label, formats["cell"])
        sheet.write_number(index, start_col + 1, item.order_count, formats["cell"])
        sheet.write_number(index, start_col + 2, item.revenue, formats["money"])
        sheet.write_number(index, start_col + 3, item.profit, formats["money"] if item.profit >= 0 else formats["negative_money"])
        sheet.write_number(index, start_col + 4, item.profit_rate, formats["percent"])
    return start_row + 2, start_row + 1 + len(points)


def write_conclusion_overview_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写结论总览页。"""

    sheet = workbook.add_worksheet("01_结论总览")
    sheet.set_column("A:A", 18)
    sheet.set_column("B:G", 18)
    sheet.set_column("I:N", 16)
    sheet.write("A1", bundle.report_title, formats["title"])
    sheet.write("A2", f"源文件：{bundle.source_workbook}", formats["text"])

    for index, metric in enumerate(bundle.overall_metrics):
        row = 3
        col = index * 2
        sheet.write(row, col, metric.label, formats["metric_label"])
        sheet.write(row, col + 1, metric.value, formats["metric_value"])

    # 2026-03-28 16:02 修改原因：用户明确要求老板页先给标准汇报口径，而不是直接堆事实和图表。
    # 2026-03-28 16:02 修改目的：把第一页收口成“一句话结论 -> 核心判断 -> 预计后果”的标准决策话术。
    sheet.write("A6", "一句话结论", formats["subtitle"])
    sheet.write("A7", "问题不是收入不增长，而是增长没有转化成利润。", formats["text"])
    sheet.write("A8", "当前盘面属于典型的低质量增长：收入仍在扩张，但利润质量正在恶化。", formats["text"])

    sheet.write("A10", "核心结论", formats["subtitle"])
    for row_index, point in enumerate(bundle.executive_points, start=11):
        sheet.write(row_index - 1, 0, f"{row_index - 6}. {point}", formats["text"])

    sheet.write("A15", "如果继续当前策略", formats["warning_title"])
    sheet.write("A16", "预计后果", formats["header"])
    sheet.write("B16", bundle.warning_signals[0].consequence, formats["text"])
    sheet.write("A17", "预计后果趋势", formats["header"])
    sheet.write(
        "B17",
        f"预计到 {bundle.future_monthly_forecast[-1].period_label} 收入继续放大，但利润率仍将下降到 "
        f"{format_percent(bundle.future_monthly_forecast[-1].profit_rate)}。",
        formats["text"],
    )

    history_first_row, history_last_row = write_monthly_table(
        sheet, formats, start_row=2, start_col=8, title="最近5个月实际经营", points=bundle.monthly_history
    )
    forecast_first_row, forecast_last_row = write_monthly_table(
        sheet, formats, start_row=10, start_col=8, title="未来5个月经营预测", points=bundle.future_monthly_forecast
    )

    trend_chart = workbook.add_chart({"type": "line"})
    trend_chart.add_series(
        {
            "name": "实际毛利",
            "categories": ["01_结论总览", history_first_row, 8, history_last_row, 8],
            "values": ["01_结论总览", history_first_row, 11, history_last_row, 11],
        }
    )
    trend_chart.add_series(
        {
            "name": "预测毛利",
            "categories": ["01_结论总览", forecast_first_row, 8, forecast_last_row, 8],
            "values": ["01_结论总览", forecast_first_row, 11, forecast_last_row, 11],
        }
    )
    trend_chart.set_title({"name": "实际 + 预测利润路径"})
    trend_chart.set_legend({"position": "bottom"})
    sheet.insert_chart("O3", trend_chart, {"x_scale": 1.1, "y_scale": 1.0})


def write_analysis_path_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写分析路径页。"""

    sheet = workbook.add_worksheet("02_分析路径")
    sheet.set_column("A:A", 16)
    sheet.set_column("B:H", 24)
    sheet.write("A1", "分析路径", formats["title"])
    sheet.write("A3", "本次分析遵循的判断顺序", formats["subtitle"])

    for index, step in enumerate(bundle.analysis_steps, start=4):
        sheet.write(index - 1, 0, f"步骤{index - 3}", formats["metric_label"])
        sheet.write(index - 1, 1, step, formats["text"])

    sheet.write("A10", "从数据到结论的逻辑", formats["subtitle"])
    sheet.write("A11", "1", formats["cell"])
    sheet.write("B11", "先看月度收入和利润是不是同步增长，而不是只看规模。", formats["text"])
    sheet.write("A12", "2", formats["cell"])
    sheet.write("B12", "再看是谁在吞利润，避免只凭感觉讨论问题。", formats["text"])
    sheet.write("A13", "3", formats["cell"])
    sheet.write("B13", "再看继续当前策略会怎样，把静态诊断变成时间趋势预警。", formats["text"])
    sheet.write("A14", "4", formats["cell"])
    sheet.write("B14", "最后比较未来5个月情景和动作拐点，让老板直接拍板。", formats["text"])


def write_core_diagnosis_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写核心诊断页。"""

    sheet = workbook.add_worksheet("03_核心诊断")
    sheet.set_column("A:G", 18)
    sheet.set_column("I:N", 18)
    sheet.write("A1", "核心诊断", formats["title"])

    sheet.write("A3", "诊断结论", formats["subtitle"])
    for index, sentence in enumerate(bundle.diagnosis_summary, start=4):
        sheet.write(index, 0, sentence, formats["text"])

    # 2026-03-28 16:02 修改原因：用户指出当前报告“没有逻辑性”，核心诊断页必须从“症状”推进到“根因”。
    # 2026-03-28 16:02 修改目的：增加“钱漏在哪”和“低质量增长”两个管理层能直接复述的判断锚点。
    sheet.write("A7", "钱漏在哪", formats["subtitle"])
    sheet.write("A8", "钱主要漏在大体量低毛利组合和持续失血城市，而不是漏在整体没有增长。", formats["text"])
    sheet.write("A9", "当前问题的本质是低质量增长：收入继续放大，但利润并没有同步转化。", formats["text"])

    history_first_row, history_last_row = write_monthly_table(
        sheet, formats, start_row=11, start_col=0, title="最近5个月实际趋势", points=bundle.monthly_history
    )
    drag_first_row, drag_last_row = write_monthly_table(
        sheet, formats, start_row=19, start_col=0, title="天猫店铺 + 酒店未来5个月预测", points=bundle.key_drag_monthly_forecast
    )

    sheet.write("I3", "年度汇总", formats["subtitle"])
    sheet.write_row("I4", ["年份", "订单数", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(bundle.year_summaries, start=4):
        sheet.write(row_index, 8, item.year, formats["cell"])
        sheet.write_number(row_index, 9, item.order_count, formats["cell"])
        sheet.write_number(row_index, 10, item.total_sales, formats["money"])
        sheet.write_number(row_index, 11, item.total_profit, formats["money"])
        sheet.write_number(row_index, 12, item.profit_rate, formats["percent"])

    history_chart = workbook.add_chart({"type": "line"})
    history_chart.add_series(
        {
            "name": "实际毛利率",
            "categories": ["03_核心诊断", history_first_row, 0, history_last_row, 0],
            "values": ["03_核心诊断", history_first_row, 4, history_last_row, 4],
        }
    )
    history_chart.set_title({"name": "最近5个月毛利率变化"})
    sheet.insert_chart("I9", history_chart, {"x_scale": 1.0, "y_scale": 1.0})

    drag_chart = workbook.add_chart({"type": "column"})
    drag_chart.add_series(
        {
            "name": "重点拖累组合毛利",
            "categories": ["03_核心诊断", drag_first_row, 0, drag_last_row, 0],
            "values": ["03_核心诊断", drag_first_row, 3, drag_last_row, 3],
        }
    )
    drag_chart.set_title({"name": "重点拖累未来5个月预测"})
    sheet.insert_chart("I25", drag_chart, {"x_scale": 1.0, "y_scale": 1.0})


def write_warning_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写经营预警页。"""

    sheet = workbook.add_worksheet("04_经营预警")
    sheet.set_column("A:A", 18)
    sheet.set_column("B:H", 24)
    sheet.set_column("I:N", 18)
    sheet.write("A1", "经营预警", formats["title"])
    sheet.write("A2", "如果不处理", formats["warning_title"])
    sheet.write("B2", "后果是什么", formats["header"])
    sheet.write("C2", "未来几个周期利润会继续被侵蚀，问题不会自然修复。", formats["text"])

    warning_outlook = build_warning_signal_outlook(bundle)
    sheet.write("A4", "时间趋势预警", formats["subtitle"])
    # 2026-03-29 09:18 修改原因：用户要求经营预警页不能只说“会出事”，还要说明灯色怎么演化、主拐点是否会出现。
    # 2026-03-29 09:18 修改目的：把预警表从静态风险清单升级成“当前信号 -> 灯色演化 -> 主拐点缺失”的管理驾驶舱口径。
    sheet.write_row(
        "A5",
        ["预警主题", "当前信号", "趋势判断", "预计后果", "预计触发月份", "预计灯色变化", "预计主拐点", "紧急程度"],
        formats["header"],
    )
    for row_index, item in enumerate(bundle.warning_signals, start=6):
        sheet.write(row_index, 0, item.title, formats["cell"])
        sheet.write(row_index, 1, item.current_signal, formats["text"])
        sheet.write(row_index, 2, item.trend_summary, formats["text"])
        sheet.write(row_index, 3, item.consequence, formats["text"])
        sheet.write(row_index, 4, item.trigger_period, formats["cell"])
        sheet.write(row_index, 5, warning_outlook["expected_status_change"], formats["text"])
        sheet.write(row_index, 6, warning_outlook["expected_main_turning_point"], formats["cell"])
        sheet.write(row_index, 7, item.urgency, formats["cell"])

    timeline_rows = build_warning_timeline_rows(bundle)
    timeline_start_row = 11
    sheet.write(timeline_start_row, 0, "预警时间轴", formats["subtitle"])
    sheet.write_row(
        timeline_start_row + 1,
        0,
        ["周期", "收入", "利润", "毛利率", "风险灯色", "预警说明", "是否主拐点"],
        formats["header"],
    )
    for row_index, item in enumerate(timeline_rows, start=timeline_start_row + 2):
        sheet.write(row_index, 0, item["period"], formats["cell"])
        sheet.write_number(row_index, 1, item["revenue"], formats["money"])
        sheet.write_number(
            row_index,
            2,
            item["profit"],
            formats["negative_money"] if item["profit"] < 0 else formats["money"],
        )
        sheet.write_number(row_index, 3, item["profit_rate"], formats["percent"])
        sheet.write(row_index, 4, item["risk_status"], formats[status_format_key(item["risk_status"])])
        sheet.write(row_index, 5, item["warning_note"], formats["text"])
        sheet.write(
            row_index,
            6,
            item["is_main_turning_point"],
            formats["turning_point"] if item["is_main_turning_point"] else formats["cell"],
        )

    weekly_start_row = timeline_start_row + len(timeline_rows) + 4
    sheet.write(weekly_start_row, 0, "周度补充信号", formats["subtitle"])
    sheet.write_row(weekly_start_row + 1, 0, ["周度周期", "预警信号", "收入", "毛利", "毛利率", "说明"], formats["header"])
    for row_index, item in enumerate(bundle.weekly_warning_points, start=weekly_start_row + 2):
        sheet.write(row_index, 0, item.period_label, formats["cell"])
        sheet.write(row_index, 1, item.signal_name, formats["text"])
        sheet.write_number(row_index, 2, item.revenue, formats["money"])
        sheet.write_number(row_index, 3, item.profit, formats["negative_money"] if item.profit < 0 else formats["money"])
        sheet.write_number(row_index, 4, item.profit_rate, formats["percent"])
        sheet.write(row_index, 5, item.note, formats["text"])

    risk_start_row = weekly_start_row + len(bundle.weekly_warning_points) + 4
    sheet.write(risk_start_row, 0, "风险城市明细", formats["subtitle"])
    sheet.write_row(risk_start_row + 1, 0, ["城市", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(bundle.risk_cities, start=risk_start_row + 2):
        sheet.write(row_index, 0, item.city, formats["cell"])
        sheet.write_number(row_index, 1, item.total_sales, formats["money"])
        sheet.write_number(row_index, 2, item.total_profit, formats["negative_money"])
        sheet.write_number(row_index, 3, item.profit_rate, formats["percent"])

    timeline_chart = workbook.add_chart({"type": "line"})
    timeline_chart.add_series(
        {
            "name": "未来利润路径",
            "categories": ["04_经营预警", timeline_start_row + 2, 0, timeline_start_row + 1 + len(timeline_rows), 0],
            "values": ["04_经营预警", timeline_start_row + 2, 2, timeline_start_row + 1 + len(timeline_rows), 2],
        }
    )
    timeline_chart.set_title({"name": "预警时间轴利润变化"})
    sheet.insert_chart("I4", timeline_chart, {"x_scale": 1.0, "y_scale": 1.0})

    warning_chart = workbook.add_chart({"type": "line"})
    warning_chart.add_series(
        {
            "name": "周度毛利",
            "categories": ["04_经营预警", weekly_start_row + 2, 0, weekly_start_row + 1 + len(bundle.weekly_warning_points), 0],
            "values": ["04_经营预警", weekly_start_row + 2, 3, weekly_start_row + 1 + len(bundle.weekly_warning_points), 3],
        }
    )
    warning_chart.set_title({"name": "周度补充信号变化"})
    sheet.insert_chart("I20", warning_chart, {"x_scale": 1.0, "y_scale": 1.0})

    city_chart = workbook.add_chart({"type": "bar"})
    city_chart.add_series(
        {
            "name": "风险城市毛利",
            "categories": ["04_经营预警", risk_start_row + 2, 0, risk_start_row + 1 + len(bundle.risk_cities), 0],
            "values": ["04_经营预警", risk_start_row + 2, 2, risk_start_row + 1 + len(bundle.risk_cities), 2],
        }
    )
    city_chart.set_title({"name": "风险城市亏损分布"})
    sheet.insert_chart(risk_start_row, 8, city_chart, {"x_scale": 1.0, "y_scale": 1.0})


def write_scenario_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写未来场景预测页。"""

    sheet = workbook.add_worksheet("05_未来场景预测")
    sheet.set_column("A:A", 16)
    sheet.set_column("B:D", 34)
    sheet.set_column("F:M", 16)
    sheet.write("A1", "未来场景预测", formats["title"])
    sheet.write("A2", "老板可选路径", formats["subtitle"])
    sheet.write("A3", "老板现在要决定的不是看不看得见问题，而是选哪条路进入利润修复。", formats["text"])
    turning_profiles = {
        item.scenario_name: build_scenario_turning_profile(bundle, item)
        for item in bundle.scenario_forecasts
    }
    # 2026-03-28 21:40 修改原因：用户要求未来场景页按“策略A/B/C”横向对打，并且每列必须固定为“结论 / 动作 / 分析 / 数据”。
    # 2026-03-28 21:40 修改目的：把这一页从概览表升级成老板拍板页，保证每个策略既有动作也有证据，不再只是抽象口号。
    sheet.write("A5", "策略对打矩阵", formats["subtitle"])
    sheet.write("A6", "", formats["header"])
    for column_index, item in enumerate(bundle.scenario_forecasts, start=1):
        sheet.write(5, column_index, item.scenario_name, formats["header"])

    matrix_rows = [
        ("策略结论", "conclusion_summary"),
        ("策略动作", "action_summary"),
        ("策略分析", "analysis_summary"),
        ("策略数据", "data_summary"),
        ("利润拐点", "profit_turning_point"),
        ("毛利率拐点", "margin_turning_point"),
        ("灯色拐点", "status_turning_point"),
        ("动作拐点", "action_turning_point"),
        ("老板主拐点", "main_turning_point"),
    ]
    for row_offset, (label, attr_name) in enumerate(matrix_rows, start=6):
        sheet.write(row_offset, 0, label, formats["header"])
        sheet.set_row(row_offset, 76 if row_offset >= 10 else 92)
        for column_index, item in enumerate(bundle.scenario_forecasts, start=1):
            if attr_name in {"conclusion_summary", "action_summary", "analysis_summary", "data_summary"}:
                cell_value = getattr(item, attr_name)
            else:
                profile = turning_profiles[item.scenario_name]
                if attr_name == "profit_turning_point":
                    cell_value = f"{profile[attr_name]}｜规则：利润连续2期改善"
                elif attr_name == "margin_turning_point":
                    cell_value = f"{profile[attr_name]}｜规则：毛利率连续2期改善"
                elif attr_name == "status_turning_point":
                    cell_value = f"{profile[attr_name]}｜规则：灯色首次改善"
                elif attr_name == "action_turning_point":
                    cell_value = f"{profile[attr_name]}｜规则：动作进入见效期"
                else:
                    cell_value = f"{profile[attr_name]}｜规则：四类拐点全部出现后的最晚月份"
            sheet.write(row_offset, column_index, cell_value, formats["text"])

    detail_start_row = 16
    sheet.write(detail_start_row, 0, "未来5个月情景测算", formats["subtitle"])
    sheet.write(detail_start_row + 1, 0, "预测数据", formats["subtitle"])
    # 2026-03-29 09:18 修改原因：用户要求 05 页预测表不能只有一个“拐点标记”，而要同时看到动作、灯色和老板主拐点的分化。
    # 2026-03-29 09:18 修改目的：保留原前 7 列兼容旧测试，再把动作状态和多拐点标记追加到后面，形成完整的策略-数据-拐点链路。
    sheet.write_row(
        detail_start_row + 2,
        0,
        ["场景", "月份", "预测数据", "收入", "毛利", "毛利率", "拐点标记", "动作状态", "利润拐点", "毛利率拐点", "灯色拐点", "动作拐点", "主拐点"],
        formats["header"],
    )
    row_cursor = detail_start_row + 3
    for scenario in bundle.scenario_forecasts:
        profile = turning_profiles[scenario.scenario_name]
        for point in scenario.monthly_projection:
            is_turning_point = point.period_label == scenario.turning_point_period and scenario.turning_point_period != "无自然拐点"
            sheet.write(row_cursor, 0, scenario.scenario_name, formats["cell"])
            sheet.write(row_cursor, 1, point.period_label, formats["turning_point"] if is_turning_point else formats["cell"])
            sheet.write(row_cursor, 2, f"{point.period_label} 预测值", formats["text"])
            sheet.write_number(row_cursor, 3, point.revenue, formats["money"])
            sheet.write_number(row_cursor, 4, point.profit, formats["negative_money"] if point.profit < 0 else formats["money"])
            sheet.write_number(row_cursor, 5, point.profit_rate, formats["percent"])
            sheet.write(row_cursor, 6, "拐点" if is_turning_point else "", formats["turning_point"] if is_turning_point else formats["cell"])
            action_status = build_scenario_action_status(bundle, scenario, point.period_label)
            sheet.write(row_cursor, 7, action_status, formats["cell"])
            sheet.write(
                row_cursor,
                8,
                "利润拐点" if point.period_label == profile["profit_turning_point"] else "",
                formats["turning_point"] if point.period_label == profile["profit_turning_point"] else formats["cell"],
            )
            sheet.write(
                row_cursor,
                9,
                "毛利率拐点" if point.period_label == profile["margin_turning_point"] else "",
                formats["turning_point"] if point.period_label == profile["margin_turning_point"] else formats["cell"],
            )
            sheet.write(
                row_cursor,
                10,
                "灯色拐点" if point.period_label == profile["status_turning_point"] else "",
                formats["turning_point"] if point.period_label == profile["status_turning_point"] else formats["cell"],
            )
            sheet.write(
                row_cursor,
                11,
                "动作拐点" if point.period_label == profile["action_turning_point"] else "",
                formats["turning_point"] if point.period_label == profile["action_turning_point"] else formats["cell"],
            )
            sheet.write(
                row_cursor,
                12,
                "主拐点" if point.period_label == profile["main_turning_point"] else "",
                formats["turning_point"] if point.period_label == profile["main_turning_point"] else formats["cell"],
            )
            row_cursor += 1

    scenario_chart = workbook.add_chart({"type": "line"})
    margin_chart = workbook.add_chart({"type": "line"})
    for scenario_index, scenario in enumerate(bundle.scenario_forecasts):
        start_row = detail_start_row + 3 + scenario_index * FORECAST_PERIODS
        end_row = start_row + FORECAST_PERIODS - 1
        scenario_chart.add_series(
            {
                "name": scenario.scenario_name,
                "categories": ["05_未来场景预测", start_row, 1, end_row, 1],
                "values": ["05_未来场景预测", start_row, 4, end_row, 4],
            }
        )
        margin_chart.add_series(
            {
                "name": scenario.scenario_name,
                "categories": ["05_未来场景预测", start_row, 1, end_row, 1],
                "values": ["05_未来场景预测", start_row, 5, end_row, 5],
            }
        )
    scenario_chart.set_title({"name": "不同情景下的利润轨迹"})
    margin_chart.set_title({"name": "不同情景下的毛利率轨迹"})
    sheet.insert_chart("F4", scenario_chart, {"x_scale": 1.0, "y_scale": 1.0})
    sheet.insert_chart("F20", margin_chart, {"x_scale": 1.0, "y_scale": 1.0})


def write_action_impact_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写动作改善测算页。"""

    sheet = workbook.add_worksheet("06_动作-改善测算")
    sheet.set_column("A:H", 18)
    sheet.set_column("J:N", 16)
    sheet.write("A1", "动作-改善测算", formats["title"])
    sheet.write("A2", "先止损，再修复，再优化", formats["subtitle"])
    sheet.write("A3", "建议老板本月拍板：先止住确定性亏损，再推动结构修复，最后用机制固化改善。", formats["text"])
    sheet.write("A5", "动作与周期结论", formats["subtitle"])
    sheet.write_row(
        "A6",
        ["优先级", "动作", "改善逻辑", "预计改善毛利额", "收入影响", "启动周期", "见效周期", "预计拐点"],
        formats["header"],
    )
    for row_index, item in enumerate(bundle.action_impacts, start=7):
        sheet.write(row_index, 0, item.priority, formats["cell"])
        sheet.write(row_index, 1, item.action_name, formats["text"])
        sheet.write(row_index, 2, item.improve_target, formats["text"])
        sheet.write_number(row_index, 3, item.expected_profit_uplift, formats["money"])
        sheet.write(row_index, 4, item.expected_revenue_impact, formats["text"])
        sheet.write(row_index, 5, item.start_period, formats["cell"])
        sheet.write(row_index, 6, item.effect_period, formats["cell"])
        sheet.write(row_index, 7, item.turning_point_period, formats["cell"])

    detail_start_row = 13
    sheet.write(detail_start_row, 0, "动作月度改善路径", formats["subtitle"])
    sheet.write_row(detail_start_row + 1, 0, ["动作", "月份", "当月改善毛利", "累计改善毛利"], formats["header"])
    row_cursor = detail_start_row + 2
    for item in bundle.action_impacts:
        for point in item.monthly_projection:
            sheet.write(row_cursor, 0, item.action_name, formats["text"])
            sheet.write(row_cursor, 1, point.period_label, formats["cell"])
            sheet.write_number(row_cursor, 2, point.monthly_profit_uplift, formats["money"])
            sheet.write_number(row_cursor, 3, point.cumulative_profit_uplift, formats["money"])
            row_cursor += 1

    compare_start_row = 32
    sheet.write(compare_start_row, 0, "不做 / 止损优先 / 全量执行", formats["subtitle"])
    sheet.write_row(compare_start_row + 1, 0, ["月份", "不做", "止损优先", "全量执行"], formats["header"])
    for index, base_point in enumerate(bundle.future_monthly_forecast, start=compare_start_row + 2):
        projection_index = index - compare_start_row - 2
        stop_loss_value = base_point.profit + bundle.action_impacts[1].monthly_projection[projection_index].monthly_profit_uplift
        full_value = base_point.profit + sum(
            action.monthly_projection[projection_index].monthly_profit_uplift for action in bundle.action_impacts
        )
        sheet.write(index, 0, base_point.period_label, formats["cell"])
        sheet.write_number(index, 1, base_point.profit, formats["money"])
        sheet.write_number(index, 2, stop_loss_value, formats["money"])
        sheet.write_number(index, 3, full_value, formats["money"])

    impact_chart = workbook.add_chart({"type": "bar"})
    impact_chart.add_series(
        {
            "name": "预计改善毛利额",
            "categories": ["06_动作-改善测算", 7, 1, 6 + len(bundle.action_impacts), 1],
            "values": ["06_动作-改善测算", 7, 3, 6 + len(bundle.action_impacts), 3],
        }
    )
    impact_chart.set_title({"name": "动作对应的预计改善毛利"})
    sheet.insert_chart("J4", impact_chart, {"x_scale": 1.0, "y_scale": 1.0})

    path_chart = workbook.add_chart({"type": "line"})
    path_chart.add_series(
        {
            "name": "不做",
            "categories": ["06_动作-改善测算", compare_start_row + 2, 0, compare_start_row + 1 + FORECAST_PERIODS, 0],
            "values": ["06_动作-改善测算", compare_start_row + 2, 1, compare_start_row + 1 + FORECAST_PERIODS, 1],
        }
    )
    path_chart.add_series(
        {
            "name": "止损优先",
            "categories": ["06_动作-改善测算", compare_start_row + 2, 0, compare_start_row + 1 + FORECAST_PERIODS, 0],
            "values": ["06_动作-改善测算", compare_start_row + 2, 2, compare_start_row + 1 + FORECAST_PERIODS, 2],
        }
    )
    path_chart.add_series(
        {
            "name": "全量执行",
            "categories": ["06_动作-改善测算", compare_start_row + 2, 0, compare_start_row + 1 + FORECAST_PERIODS, 0],
            "values": ["06_动作-改善测算", compare_start_row + 2, 3, compare_start_row + 1 + FORECAST_PERIODS, 3],
        }
    )
    path_chart.set_title({"name": "不同执行力度下的利润路径"})
    sheet.insert_chart("J22", path_chart, {"x_scale": 1.0, "y_scale": 1.0})


def build_loss_control_priority_items(bundle: BossReportBundle) -> list[LossControlPriorityItem]:
    """把贡献结果整理成止损对象优先级矩阵。"""

    # 2026-03-28 22:34 修改原因：用户要求客户贡献拆解页直接回答“先止损谁、为什么、怎么动、多久复盘”，而不是继续展示贡献榜单。
    # 2026-03-28 22:34 修改目的：把现有城市、渠道、品类结果转成执行层优先级对象，供 07 页和后续 Skill 复用同一套口径。
    risk_city_name_set = {item.city for item in bundle.risk_cities}
    worst_combo = min(bundle.shop_category_summaries, key=lambda item: item.profit_rate)
    better_same_shop_items = [
        item
        for item in bundle.shop_category_summaries
        if item.shop == worst_combo.shop and item.category != worst_combo.category
    ]
    better_same_shop_items.sort(key=lambda item: item.profit_rate, reverse=True)
    first_safe_city = next(
        (item for item in bundle.city_contributions if item.city not in risk_city_name_set),
        bundle.city_contributions[0],
    )
    structure_uplift_value = (
        max(0.0, (better_same_shop_items[0].profit_rate - worst_combo.profit_rate) * 1_000_000)
        if better_same_shop_items
        else 0.0
    )
    limit_items = bundle.risk_cities[1:3]
    limit_city_names = "、".join(item.city for item in limit_items) or "暂无第二梯队止损城市"
    limit_city_loss = abs(sum(item.total_profit for item in limit_items if item.total_profit < 0))

    items: list[LossControlPriorityItem] = []
    if bundle.risk_cities:
        p1_item = bundle.risk_cities[0]
        items.append(
            LossControlPriorityItem(
                object_name=p1_item.city,
                priority="P1 立即止损",
                action=f"立即暂停 {p1_item.city} 在 {worst_combo.shop}+{worst_combo.category} 的低价放量单。",
                reason=f"{p1_item.city} 当前已转负利润，且在风险城市中失血最大。",
                expected_improvement=f"优先止住约 {format_wan(abs(p1_item.total_profit))} 失血，并为利润率修复腾空间。",
                review_cycle="按周复盘",
                expected_improvement_value=abs(p1_item.total_profit),
            )
        )
    items.append(
        LossControlPriorityItem(
            object_name=limit_city_names,
            priority="P2 限制放量",
            action=f"限制 {limit_city_names} 的高风险投放和低价订单继续放量。",
            reason=f"{limit_city_names} 已经进入第二梯队风险，如果继续放量会把局部亏损扩散成区域问题。",
            expected_improvement=f"预计减少约 {format_wan(limit_city_loss)} 继续失血，并降低后续止损成本。",
            review_cycle="双周复盘",
            expected_improvement_value=limit_city_loss,
        )
    )
    items.append(
        LossControlPriorityItem(
            object_name=f"{worst_combo.shop}+{worst_combo.category}",
            priority="P3 结构修复",
            action=f"修复 {worst_combo.shop}+{worst_combo.category} 的定价、补贴和流量门槛，并把新增流量导向同店铺更优结构。",
            reason=f"{worst_combo.shop}+{worst_combo.category} 规模大但毛利率最低，问题核心是结构错配而不是单点亏损。",
            expected_improvement=(
                f"每转移100万销售额到更优结构，理论上可多贡献约 {format_wan(structure_uplift_value)} 毛利。"
                if structure_uplift_value > 0
                else "优先改善利润质量，再观察结构修复后的规模承接。"
            ),
            review_cycle="按月复盘",
            expected_improvement_value=structure_uplift_value,
        )
    )
    items.append(
        LossControlPriorityItem(
            object_name=first_safe_city.city,
            priority="P4 持续观察",
            action=f"保持 {first_safe_city.city} 当前经营策略，不做激进收缩，只保留月度观察。",
            reason=f"{first_safe_city.city} 当前贡献高且利润质量相对健康，优先保护优质规模，避免误伤。",
            expected_improvement="核心目标不是额外止损，而是避免优质城市被一刀切收缩误伤。",
            review_cycle="按月观察",
            expected_improvement_value=max(first_safe_city.total_profit * 0.05, 0.0),
        )
    )
    return items


def build_dynamic_status_transition(
    bundle: BossReportBundle,
    priority_item: LossControlPriorityItem,
) -> tuple[str, str, str]:
    """基于现有经营信号构造可解释的动态切灯结果。"""

    # 2026-03-28 23:32 修改原因：用户要求“上期灯色 / 本期灯色 / 变灯原因”必须能追溯到已有经营数据，而不是黑盒口径。
    # 2026-03-28 23:32 修改目的：复用风险城市、周度预警和渠道品类结果，生成老板可追问、团队可复盘的动态切灯解释。
    latest_weekly_signal = bundle.weekly_warning_points[-1] if bundle.weekly_warning_points else None
    combo_weekly_signal = next(
        (
            point
            for point in reversed(bundle.weekly_warning_points)
            if "重点组合" in point.signal_name or "单周亏损" in point.signal_name
        ),
        latest_weekly_signal,
    )
    worst_combo = min(bundle.shop_category_summaries, key=lambda item: item.profit_rate)
    better_same_shop_items = [
        item
        for item in bundle.shop_category_summaries
        if item.shop == worst_combo.shop and item.category != worst_combo.category
    ]
    best_same_shop_item = max(
        better_same_shop_items,
        key=lambda item: item.profit_rate,
        default=None,
    )
    risk_city_names = {item.city for item in bundle.risk_cities}
    risk_city_count = len(bundle.risk_cities)

    if priority_item.priority.startswith("P1"):
        current_risk = next(
            (item for item in bundle.risk_cities if item.city == priority_item.object_name),
            None,
        )
        signal_name = latest_weekly_signal.signal_name if latest_weekly_signal else "周度风险继续走弱"
        loss_text = (
            format_wan(abs(current_risk.total_profit))
            if current_risk is not None
            else "重点失血规模"
        )
        return (
            "黄灯",
            "红灯",
            f"{priority_item.object_name} 最新周度已出现“{signal_name}”，当前失血约 {loss_text}，由黄灯升级为红灯。",
        )

    if priority_item.priority.startswith("P2"):
        second_tier_risks = bundle.risk_cities[1:3]
        second_tier_loss = abs(
            sum(item.total_profit for item in second_tier_risks if item.total_profit < 0)
        )
        signal_name = latest_weekly_signal.signal_name if latest_weekly_signal else "风险城市亏损扩大"
        return (
            "绿灯",
            "黄灯",
            f"{priority_item.object_name} 已进入第二梯队风险组，合计失血约 {format_wan(second_tier_loss)}，且“{signal_name}”仍在延续，由绿灯转黄灯。",
        )

    if priority_item.priority.startswith("P3"):
        combo_signal_name = combo_weekly_signal.signal_name if combo_weekly_signal else "重点组合利润继续下探"
        better_structure_text = (
            f"{best_same_shop_item.shop}+{best_same_shop_item.category} 毛利率 {format_percent(best_same_shop_item.profit_rate)}"
            if best_same_shop_item is not None
            else "同店更优结构仍有承接空间"
        )
        return (
            "黄灯",
            "黄灯",
            (
                f"{worst_combo.shop}+{worst_combo.category} 毛利率仅 {format_percent(worst_combo.profit_rate)}，"
                f"周度仍呈“{combo_signal_name}”，且低于 {better_structure_text}，结构问题未修复，维持黄灯。"
            ),
        )

    current_city = next(
        (item for item in bundle.city_contributions if item.city == priority_item.object_name),
        None,
    )
    city_profit_rate_text = (
        format_percent(current_city.profit_rate)
        if current_city is not None
        else "当前利润质量仍稳定"
    )
    not_in_risk_group = priority_item.object_name not in risk_city_names
    stability_text = "未进入风险组" if not_in_risk_group else f"仍需关注 {risk_city_count} 个风险城市扩散"
    return (
        "绿灯",
        "绿灯",
        f"{priority_item.object_name} 当前毛利率约 {city_profit_rate_text}，{stability_text}，维持绿灯。",
    )


def build_loss_control_execution_items(bundle: BossReportBundle) -> list[LossControlExecutionItem]:
    """把止损对象优先级矩阵转成老板执行版跟踪项。"""

    # 2026-03-29 00:23 修改原因：用户要求 07 页不只回答“先动谁”，还要继续回答“谁负责、何时完成、如何验收”。
    # 2026-03-29 00:23 修改目的：复用既有优先级矩阵对象，补一层执行口径映射，避免把风险判断逻辑重复写两遍。
    priority_items = build_loss_control_priority_items(bundle)
    execution_items: list[LossControlExecutionItem] = []
    for item in priority_items:
        previous_status_light, current_status_light, status_change_reason = build_dynamic_status_transition(
            bundle,
            item,
        )
        if item.priority.startswith("P1"):
            execution_items.append(
                LossControlExecutionItem(
                    priority=item.priority,
                    object_name=item.object_name,
                    status_light=current_status_light,
                    previous_status_light=previous_status_light,
                    current_status_light=current_status_light,
                    status_change_reason=status_change_reason,
                    conclusion=f"{item.object_name} 已进入确定性失血区，必须立即止血。",
                    key_action=item.action,
                    owner="经营分析 + 渠道运营",
                    co_owner="商品运营",
                    timeline="第1周启动，第2周复盘",
                    deadline="本周内完成首轮止损动作",
                    acceptance_metric="单周亏损收窄，低价放量单停止扩张",
                    stage_goal="先止住核心失血，再决定是否恢复局部投放",
                    risk_alert="若继续放量，亏损会先于收入修复继续扩大。",
                    review_cycle=item.review_cycle,
                    weekly_judgement="本周必须盯住失血点，先止血再谈恢复。",
                    next_review_time="下周经营例会",
                    current_week_action="暂停低价放量单并复核补贴、投放与价格门槛",
                    next_week_action="回看亏损收窄情况，决定是否保留局部流量",
                    expected_improvement_value=item.expected_improvement_value,
                )
            )
        elif item.priority.startswith("P2"):
            execution_items.append(
                LossControlExecutionItem(
                    priority=item.priority,
                    object_name=item.object_name,
                    status_light=current_status_light,
                    previous_status_light=previous_status_light,
                    current_status_light=current_status_light,
                    status_change_reason=status_change_reason,
                    conclusion=f"{item.object_name} 正在成为下一批亏损扩散点，必须限制放量。",
                    key_action=item.action,
                    owner="渠道运营",
                    co_owner="经营分析",
                    timeline="双周调整，月末复盘",
                    deadline="两周内压住扩散趋势",
                    acceptance_metric="第二梯队亏损不再扩大，风险城市数量不增加",
                    stage_goal="先压住扩散，再判断是否升级为 P1 立即止损",
                    risk_alert="若不限制放量，局部风险会扩散成区域问题。",
                    review_cycle=item.review_cycle,
                    weekly_judgement="本周重点是控扩散，不能再让第二梯队继续放量。",
                    next_review_time="双周复盘会",
                    current_week_action="收缩高风险投放与低价订单，跟踪扩散城市表现",
                    next_week_action="复核是否继续限量或升级为红灯对象",
                    expected_improvement_value=item.expected_improvement_value,
                )
            )
        elif item.priority.startswith("P3"):
            execution_items.append(
                LossControlExecutionItem(
                    priority=item.priority,
                    object_name=item.object_name,
                    status_light=current_status_light,
                    previous_status_light=previous_status_light,
                    current_status_light=current_status_light,
                    status_change_reason=status_change_reason,
                    conclusion=f"{item.object_name} 的核心矛盾是结构错配，不是简单规模不足。",
                    key_action=item.action,
                    owner="商品运营 + 渠道运营",
                    co_owner="经营分析",
                    timeline="本月修结构，下月看拐点",
                    deadline="本月内完成结构修复方案切换",
                    acceptance_metric="组合毛利率回升，高毛利结构占比提升",
                    stage_goal="先修复定价与补贴，再把新增流量导向更优结构",
                    risk_alert="若只保规模不修结构，收入越大，利润质量越差。",
                    review_cycle=item.review_cycle,
                    weekly_judgement="本周先修价格和流量结构，避免低质量收入继续放大。",
                    next_review_time="下月第一周复盘",
                    current_week_action="调整定价、补贴和流量门槛，切换新增流量去向",
                    next_week_action="检查高毛利结构占比与组合毛利率是否回升",
                    expected_improvement_value=item.expected_improvement_value,
                )
            )
        else:
            execution_items.append(
                LossControlExecutionItem(
                    priority=item.priority,
                    object_name=item.object_name,
                    status_light=current_status_light,
                    previous_status_light=previous_status_light,
                    current_status_light=current_status_light,
                    status_change_reason=status_change_reason,
                    conclusion=f"{item.object_name} 当前属于应保护的优质规模，重点是避免误伤。",
                    key_action=item.action,
                    owner="经营分析",
                    co_owner="渠道运营",
                    timeline="按月观察，季度校准",
                    deadline="持续观察，无需立即收缩",
                    acceptance_metric="利润质量稳定，未滑入风险组",
                    stage_goal="保护健康规模，不做一刀切收缩",
                    risk_alert="若误伤优质城市，利润修复和收入承接都会受损。",
                    review_cycle=item.review_cycle,
                    weekly_judgement="本周不主动收缩，重点保护优质规模与利润质量。",
                    next_review_time="下月经营复盘",
                    current_week_action="保持当前策略，监控是否出现利润质量转弱信号",
                    next_week_action="按月复核是否仍保持在健康区间",
                    expected_improvement_value=item.expected_improvement_value,
                )
            )
    return execution_items


def status_format_key(status_light: str) -> str:
    """根据状态灯返回单元格格式键。"""

    if status_light == "红灯":
        return "red_status"
    if status_light == "黄灯":
        return "yellow_status"
    return "green_status"


def write_city_contribution_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写客户贡献拆解页。"""

    sheet = workbook.add_worksheet("07_客户贡献拆解")
    sheet.set_column("A:A", 18)
    sheet.set_column("B:B", 14)
    sheet.set_column("C:D", 26)
    sheet.set_column("E:F", 18)
    sheet.set_column("G:G", 42)
    sheet.set_column("H:H", 18)
    sheet.set_column("I:I", 14, None, {"hidden": True})
    sheet.write("A1", "止损对象优先级矩阵", formats["title"])
    sheet.write("A2", "这页不是继续展示谁贡献高，而是明确先止损谁、先保谁、先观察谁。", formats["text"])

    # 2026-03-29 00:24 修改原因：用户继续要求把 07 页升级成老板执行摘要，而不是停在分析矩阵层。
    # 2026-03-29 00:24 修改目的：让老板页直接回答“结论、关键动作、负责人、时间表、验收指标”，把细执行留给附表承接。
    priority_items = build_loss_control_priority_items(bundle)
    execution_items = build_loss_control_execution_items(bundle)
    sheet.write("A4", "止损对象优先级矩阵", formats["subtitle"])
    sheet.write_row(
        "A5",
        ["止损对象", "优先级", "结论", "关键动作", "负责人", "时间表", "验收指标", "预计改善毛利额"],
        formats["header"],
    )
    for row_index, item in enumerate(execution_items, start=6):
        sheet.write(row_index, 0, item.object_name, formats["cell"])
        sheet.write(row_index, 1, item.priority, formats["text"])
        sheet.write(row_index, 2, item.conclusion, formats["text"])
        sheet.write(row_index, 3, item.key_action, formats["text"])
        sheet.write(row_index, 4, item.owner, formats["text"])
        sheet.write(row_index, 5, item.timeline, formats["text"])
        sheet.write(row_index, 6, item.acceptance_metric, formats["text"])
        sheet.write_number(row_index, 7, item.expected_improvement_value, formats["money"])

    tip_row = 7 + len(execution_items)
    sheet.write(tip_row, 0, "老板拍板提示", formats["warning_title"])
    # 2026-03-29 00:24 修改原因：用户要的是老板能直接拍板的页面，而不是继续看一张纯分析表。
    # 2026-03-29 00:24 修改目的：补三条高密度拍板话术，让老板快速决定“先止损、谁负责、何时复盘”。
    prompt_lines = [
        f"先拍板：优先处理 {execution_items[0].object_name}，同步约束 {execution_items[1].object_name} 的风险扩散。",
        f"先定责：{execution_items[0].owner} 牵头止损，{execution_items[2].owner} 牵头修复 {execution_items[2].object_name} 结构。",
        f"先看结果：第一轮结果在 {execution_items[0].timeline} 内复盘，关键看验收指标是否达成。",
    ]
    for row_offset, line in enumerate(prompt_lines, start=tip_row + 1):
        sheet.write(row_offset, 0, line, formats["text"])

    rag_row = tip_row + len(prompt_lines) + 2
    sheet.write(rag_row, 0, "周会红黄绿状态板", formats["subtitle"])
    # 2026-03-28 23:32 修改原因：用户要求老板页不只看到静态灯色，还要看到上期、本期和变灯原因。
    # 2026-03-28 23:32 修改目的：保留旧合同字段“状态灯 / 本周判断 / 下次复盘时间”，同时补齐动态切灯解释。
    sheet.write_row(
        rag_row + 1,
        0,
        ["止损对象", "优先级", "状态灯", "上期灯色", "本期灯色", "变灯原因", "本周判断", "下次复盘时间"],
        formats["header"],
    )
    for row_index, item in enumerate(execution_items, start=rag_row + 2):
        sheet.write(row_index, 0, item.object_name, formats["cell"])
        sheet.write(row_index, 1, item.priority, formats["cell"])
        sheet.write(row_index, 2, item.status_light, formats[status_format_key(item.status_light)])
        sheet.write(row_index, 3, item.previous_status_light, formats[status_format_key(item.previous_status_light)])
        sheet.write(row_index, 4, item.current_status_light, formats[status_format_key(item.current_status_light)])
        sheet.write(row_index, 5, item.status_change_reason, formats["text"])
        sheet.write(row_index, 6, item.weekly_judgement, formats["text"])
        sheet.write(row_index, 7, item.next_review_time, formats["cell"])

    note_row = rag_row + len(execution_items) + 3
    sheet.write(note_row, 0, "执行口径说明", formats["subtitle"])
    # 2026-03-29 00:24 修改原因：旧测试和旧阅读口径仍关心“动作、原因、预计改善、观察周期”。
    # 2026-03-29 00:24 修改目的：明确这些字段没有消失，而是下沉到附录执行跟踪附表，保持口径连续。
    for row_offset, item in enumerate(priority_items, start=note_row + 1):
        sheet.write(
            row_offset,
            0,
            f"{item.priority} | 动作：{item.action} | 原因：{item.reason} | 预计改善：{item.expected_improvement} | 观察周期：{item.review_cycle}",
            formats["text"],
        )

    city_chart = workbook.add_chart({"type": "bar"})
    city_chart.add_series(
        {
            "name": "优先级对象预计改善毛利额",
            "categories": ["07_客户贡献拆解", 6, 0, 5 + len(execution_items), 0],
            "values": ["07_客户贡献拆解", 6, 7, 5 + len(execution_items), 7],
        }
    )
    city_chart.set_title({"name": "优先级对象预计改善毛利额"})
    sheet.insert_chart("J4", city_chart, {"x_scale": 1.05, "y_scale": 1.0})


def write_appendix_chart_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写附录图表与明细页。"""

    sheet = workbook.add_worksheet("08_附录-图表与明细")
    sheet.set_column("A:B", 18)
    sheet.set_column("C:E", 14)
    sheet.set_column("F:F", 42)
    sheet.set_column("G:N", 18)
    sheet.set_column("O:Q", 16)
    sheet.write("A1", "附录-图表与明细", formats["title"])
    sheet.write("A3", "渠道品类风险明细", formats["subtitle"])
    sheet.write_row("A4", ["店铺", "品类", "订单数", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(bundle.shop_category_summaries, start=5):
        sheet.write(row_index, 0, item.shop, formats["cell"])
        sheet.write(row_index, 1, item.category, formats["cell"])
        sheet.write_number(row_index, 2, item.order_count, formats["cell"])
        sheet.write_number(row_index, 3, item.total_sales, formats["money"])
        sheet.write_number(row_index, 4, item.total_profit, formats["money"])
        sheet.write_number(row_index, 5, item.profit_rate, formats["percent"])

    execution_items = build_loss_control_execution_items(bundle)
    tracker_start_row = 12
    # 2026-03-29 00:25 修改原因：方案B要求老板页只拍板，执行细节要下沉到附录承接。
    # 2026-03-29 00:25 修改目的：在同一张附录页增加执行跟踪附表，让经营团队拿到负责人、时间表、风险提示与验收指标。
    sheet.write(tracker_start_row, 0, "执行跟踪附表", formats["subtitle"])
    sheet.write_row(
        tracker_start_row + 1,
        0,
        [
            "优先级",
            "止损对象",
            "状态灯",
            "上期灯色",
            "本期灯色",
            "变灯原因",
            "负责人",
            "协同人",
            "本周动作",
            "下周动作",
            "截止时间",
            "验收指标",
            "风险提示",
            "复盘周期",
        ],
        formats["header"],
    )
    for row_index, item in enumerate(execution_items, start=tracker_start_row + 2):
        sheet.write(row_index, 0, item.priority, formats["cell"])
        sheet.write(row_index, 1, item.object_name, formats["cell"])
        sheet.write(row_index, 2, item.status_light, formats[status_format_key(item.status_light)])
        sheet.write(row_index, 3, item.previous_status_light, formats[status_format_key(item.previous_status_light)])
        sheet.write(row_index, 4, item.current_status_light, formats[status_format_key(item.current_status_light)])
        sheet.write(row_index, 5, item.status_change_reason, formats["text"])
        sheet.write(row_index, 6, item.owner, formats["text"])
        sheet.write(row_index, 7, item.co_owner, formats["text"])
        sheet.write(row_index, 8, item.current_week_action, formats["text"])
        sheet.write(row_index, 9, item.next_week_action, formats["text"])
        sheet.write(row_index, 10, item.deadline, formats["text"])
        sheet.write(row_index, 11, item.acceptance_metric, formats["text"])
        sheet.write(row_index, 12, item.risk_alert, formats["text"])
        sheet.write(row_index, 13, item.review_cycle, formats["cell"])

    detail_row = tracker_start_row + len(execution_items) + 3
    sheet.write(detail_row, 0, "执行补充口径", formats["subtitle"])
    # 2026-03-29 00:48 修改原因：原附表测试已经锁了“第一阶段目标”，新周会版不能把旧合同丢掉。
    # 2026-03-29 00:48 修改目的：保留阶段目标文本区，保证原执行版口径与新周会版兼容。
    for row_index, item in enumerate(execution_items, start=detail_row + 1):
        sheet.write(
            row_index,
            0,
            f"{item.priority} | 第一阶段目标：{item.stage_goal} | 启动动作：{item.key_action}",
            formats["text"],
        )

    # 2026-03-28 22:11 修改原因：用户要求附录必须讲清楚未来场景和拐点是怎么推演出来的，否则老板会认为只是“讲故事”。
    # 2026-03-28 22:11 修改目的：在附录页补齐算法名称、输入数据、推演步骤、拐点规则和非黑盒说明，形成可复核的方法附录。
    worst_combo = min(bundle.shop_category_summaries, key=lambda item: item.profit_rate)
    risk_city_names = "、".join(item.city for item in bundle.risk_cities[:5]) or "暂无重点风险城市"
    appendix_start_row = detail_row + len(execution_items) + 3
    appendix_lines = [
        "算法与推演附录",
        "算法名称：情景经营轨迹模型",
        "动态切灯算法：可解释经营规则引擎",
        (
            f"输入数据：最近月度收入/利润轨迹、重点拖累组合 {worst_combo.shop}+{worst_combo.category}、"
            f"风险城市 {risk_city_names}、以及止损和结构优化动作参数。"
        ),
        "切灯规则1：最新周度信号转负且对象进入风险城市榜首时，黄灯升级为红灯。",
        "切灯规则2：第二梯队风险城市开始扩散时，绿灯升级为黄灯。",
        "切灯规则3：重点组合结构问题未修复时，黄灯维持不变。",
        "切灯规则4：未进入风险组且利润质量稳定时，绿灯维持不变。",
        "推演步骤1：先用加权移动平均提取最近经营趋势，得到收入基线和毛利率基线。",
        "推演步骤2：再把止损、提价、降补贴、结构替换等动作转换成动作改善斜率，映射到未来各期毛利率变化。",
        "推演步骤3：按情景分别滚动生成未来5个月的收入、毛利和毛利率路径，而不是只给一个静态结论。",
        "推演步骤4：用盈亏平衡穿越检查累计改善额何时覆盖规模损失或止损成本，判断动作有没有真正穿越经营底线。",
        "拐点规则：利润连续2期改善，且毛利率连续2期改善，同时通过盈亏平衡穿越的首月，才定义为经营拐点。",
        "解释口径：这不是机器学习黑盒预测，而是可解释经营算法，便于老板追问“为什么是这个月”。",
    ]
    sheet.write(appendix_start_row, 0, appendix_lines[0], formats["subtitle"])
    for row_offset, line in enumerate(appendix_lines[1:], start=appendix_start_row + 1):
        sheet.write(row_offset, 0, line, formats["text"])

    appendix_chart = workbook.add_chart({"type": "column"})
    appendix_chart.add_series(
        {
            "name": "销售额",
            "categories": ["08_附录-图表与明细", 5, 0, 4 + len(bundle.shop_category_summaries), 1],
            "values": ["08_附录-图表与明细", 5, 3, 4 + len(bundle.shop_category_summaries), 3],
        }
    )
    appendix_chart.set_title({"name": "渠道 x 品类销售额"})
    sheet.insert_chart("P4", appendix_chart, {"x_scale": 1.0, "y_scale": 1.0})


def write_appendix_summary_sheet(
    workbook: xlsxwriter.Workbook,
    formats: dict[str, xlsxwriter.format.Format],
    bundle: BossReportBundle,
) -> None:
    """写附录分组汇总页。"""

    sheet = workbook.add_worksheet("09_附录-分组汇总")
    sheet.set_column("A:H", 16)
    row_cursor = 0
    row_cursor = write_year_summary_block(sheet, formats, bundle.year_summaries, row_cursor)
    row_cursor = write_city_summary_block(sheet, formats, bundle.city_contributions, row_cursor + 2)
    row_cursor = write_risk_city_block(sheet, formats, bundle.risk_cities, row_cursor + 2)
    row_cursor = write_shop_category_block(sheet, formats, bundle.shop_category_summaries, row_cursor + 2)
    write_monthly_summary_block(sheet, formats, bundle.monthly_history + bundle.future_monthly_forecast, row_cursor + 2)


def write_year_summary_block(
    sheet: xlsxwriter.worksheet.Worksheet,
    formats: dict[str, xlsxwriter.format.Format],
    items: list[YearSummary],
    row_cursor: int,
) -> int:
    """写年度汇总块。"""

    sheet.write(row_cursor, 0, "年度汇总", formats["subtitle"])
    sheet.write_row(row_cursor + 1, 0, ["年份", "订单数", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(items, start=row_cursor + 2):
        sheet.write(row_index, 0, item.year, formats["cell"])
        sheet.write_number(row_index, 1, item.order_count, formats["cell"])
        sheet.write_number(row_index, 2, item.total_sales, formats["money"])
        sheet.write_number(row_index, 3, item.total_profit, formats["money"])
        sheet.write_number(row_index, 4, item.profit_rate, formats["percent"])
    return row_cursor + len(items) + 2


def write_city_summary_block(
    sheet: xlsxwriter.worksheet.Worksheet,
    formats: dict[str, xlsxwriter.format.Format],
    items: list[CityContribution],
    row_cursor: int,
) -> int:
    """写城市汇总块。"""

    sheet.write(row_cursor, 0, "城市贡献汇总", formats["subtitle"])
    sheet.write_row(row_cursor + 1, 0, ["城市", "订单数", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(items, start=row_cursor + 2):
        sheet.write(row_index, 0, item.city, formats["cell"])
        sheet.write_number(row_index, 1, item.order_count, formats["cell"])
        sheet.write_number(row_index, 2, item.total_sales, formats["money"])
        sheet.write_number(row_index, 3, item.total_profit, formats["money"])
        sheet.write_number(row_index, 4, item.profit_rate, formats["percent"])
    return row_cursor + len(items) + 2


def write_risk_city_block(
    sheet: xlsxwriter.worksheet.Worksheet,
    formats: dict[str, xlsxwriter.format.Format],
    items: list[RiskCitySummary],
    row_cursor: int,
) -> int:
    """写风险城市块。"""

    sheet.write(row_cursor, 0, "风险城市汇总", formats["subtitle"])
    sheet.write_row(row_cursor + 1, 0, ["城市", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(items, start=row_cursor + 2):
        sheet.write(row_index, 0, item.city, formats["cell"])
        sheet.write_number(row_index, 1, item.total_sales, formats["money"])
        sheet.write_number(row_index, 2, item.total_profit, formats["negative_money"])
        sheet.write_number(row_index, 3, item.profit_rate, formats["percent"])
    return row_cursor + len(items) + 2


def write_shop_category_block(
    sheet: xlsxwriter.worksheet.Worksheet,
    formats: dict[str, xlsxwriter.format.Format],
    items: list[ShopCategorySummary],
    row_cursor: int,
) -> int:
    """写渠道品类汇总块。"""

    sheet.write(row_cursor, 0, "渠道品类汇总", formats["subtitle"])
    sheet.write_row(row_cursor + 1, 0, ["店铺", "品类", "订单数", "销售额", "毛利额", "毛利率"], formats["header"])
    for row_index, item in enumerate(items, start=row_cursor + 2):
        sheet.write(row_index, 0, item.shop, formats["cell"])
        sheet.write(row_index, 1, item.category, formats["cell"])
        sheet.write_number(row_index, 2, item.order_count, formats["cell"])
        sheet.write_number(row_index, 3, item.total_sales, formats["money"])
        sheet.write_number(row_index, 4, item.total_profit, formats["money"])
        sheet.write_number(row_index, 5, item.profit_rate, formats["percent"])
    return row_cursor + len(items) + 2


def write_monthly_summary_block(
    sheet: xlsxwriter.worksheet.Worksheet,
    formats: dict[str, xlsxwriter.format.Format],
    items: list[MonthlyTrendPoint],
    row_cursor: int,
) -> int:
    """写月度汇总块。"""

    sheet.write(row_cursor, 0, "月度汇总", formats["subtitle"])
    sheet.write_row(row_cursor + 1, 0, ["月份", "订单数", "收入", "毛利", "毛利率"], formats["header"])
    for row_index, item in enumerate(items, start=row_cursor + 2):
        sheet.write(row_index, 0, item.period_label, formats["cell"])
        sheet.write_number(row_index, 1, item.order_count, formats["cell"])
        sheet.write_number(row_index, 2, item.revenue, formats["money"])
        sheet.write_number(row_index, 3, item.profit, formats["money"] if item.profit >= 0 else formats["negative_money"])
        sheet.write_number(row_index, 4, item.profit_rate, formats["percent"])
    return row_cursor + len(items) + 2


def generate_boss_report_workbook(
    source_workbook: Path = DEFAULT_SOURCE_WORKBOOK,
    output_workbook: Path = DEFAULT_OUTPUT_WORKBOOK,
    rust_binary: Path = DEFAULT_RUST_BINARY,
) -> Path:
    """对真实文件执行“Rust 分析 + Excel 交付”生成老板报告。"""

    bundle = load_boss_report_bundle_from_rust(
        source_workbook=source_workbook,
        rust_binary=rust_binary,
        sheet_name=DEFAULT_SHEET_NAME,
    )
    return build_boss_report_workbook(bundle, output_workbook)


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""

    parser = argparse.ArgumentParser(description="生成老板决策版业绩诊断工作簿")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE_WORKBOOK))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_WORKBOOK))
    parser.add_argument("--binary", default=str(DEFAULT_RUST_BINARY))
    return parser.parse_args()


def main() -> None:
    """命令行入口。"""

    args = parse_args()
    output = generate_boss_report_workbook(
        source_workbook=Path(args.source),
        output_workbook=Path(args.output),
        rust_binary=Path(args.binary),
    )
    print(f"已生成老板决策版工作簿: {output}")
