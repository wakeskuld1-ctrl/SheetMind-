# -*- coding: utf-8 -*-

"""Excel 脱敏副本生成工具。

这个脚本的职责很单一：
1. 复制指定的 Excel 文件到目标目录。
2. 仅改写目标 sheet 的数据区，不改 sheet 名、表头、基础结构。
3. 用符合保险业务语境、带淡旺季波动的假数据替换原始数据。
"""

from __future__ import annotations

import argparse
import math
import shutil
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

try:
    import pythoncom
    import win32com.client as win32
except ModuleNotFoundError:  # pragma: no cover - 这里只做运行时能力探测，测试不强依赖 COM。
    pythoncom = None
    win32 = None


# 2026-03-22 13:46 CST: 修改原因和目的：把真实交付需求固化成默认任务配置，避免运行时再手工拼路径导致误处理客户文件。
DEFAULT_JOBS = (
    {
        "source_path": Path(r"D:\Excel测试\新疆客户\直保业务收入明细表-2025全部收入.xlsx"),
        "output_name": "澄岳保险集团-产险事业群-2025经营收入总台账.xlsx",
        "target_sheet_names": None,
        "workbook_kind": "property_2025",
    },
    {
        "source_path": Path(r"D:\Excel测试\新疆客户\2026文旅体台账.xlsx"),
        "output_name": "澄岳保险集团-寿险事业群-2026业务经营总台账.xlsx",
        "target_sheet_names": None,
        "workbook_kind": "life_2026",
    },
    {
        "source_path": Path(r"D:\Excel测试\新疆客户\数据处理器.xlsm"),
        "output_name": "澄岳保险集团-经营管理中台-业务数据处理器.xlsm",
        "target_sheet_names": {"客户信息"},
        "workbook_kind": "ops_center",
    },
)


# 2026-03-22 13:46 CST: 修改原因和目的：集中维护固定词库，让假数据既统一又容易调整，避免在多处分散硬编码。
INSURER_PREFIXES = ("远宸", "衡泰", "华澜", "安岳", "云晟", "恒祺", "嘉禾", "启明", "天裕", "瑞衡")
INSURER_SUFFIXES = ("保险股份有限公司", "财产保险股份有限公司", "人寿保险股份有限公司", "保险服务有限公司")
CUSTOMER_PREFIXES = ("星河", "云途", "盛景", "华旅", "安程", "嘉境", "明远", "恒信", "天域", "锦川")
CUSTOMER_SUFFIXES = (
    "文旅发展有限公司",
    "商贸有限公司",
    "供应链管理有限公司",
    "健康管理有限公司",
    "汽车服务有限公司",
    "物业服务有限公司",
    "科技有限公司",
    "农业发展有限公司",
    "工业设备有限公司",
    "会展服务有限公司",
)
PERSON_NAMES = (
    "周承泽",
    "林知夏",
    "高昱辰",
    "顾清和",
    "唐予安",
    "许嘉木",
    "宋昭宁",
    "韩景川",
    "叶书禾",
    "陆闻笙",
)
# 2026-03-22 15:55 CST: 修改原因和目的：将地域池从新疆本地城市改为全国核心城市混合，避免继续暴露单一区域指向。
REGIONS = (
    "北京",
    "上海",
    "广州",
    "深圳",
    "杭州",
    "南京",
    "苏州",
    "成都",
    "重庆",
    "武汉",
    "长沙",
    "郑州",
    "西安",
    "青岛",
    "厦门",
    "宁波",
    "天津",
    "福州",
)
DEPARTMENTS = (
    "综合金融事业部",
    "企业客户一部",
    "产险渠道部",
    "寿险顾问部",
    "重点客户经营部",
    "车险拓展部",
    "责任险项目组",
)
BUSINESS_CHANNELS = ("渠道", "互联网", "直客", "传统经纪", "联合展业")
PRIMARY_CATEGORIES = ("财产", "寿险", "健康险", "责任险", "车险")
SECONDARY_CATEGORIES = ("渠道", "互联网", "传统经纪", "团险", "企财", "责任险")
INSURANCE_PRODUCTS = (
    "雇主责任险",
    "企财险",
    "公众责任险",
    "团体意外险",
    "建筑工程险",
    "定期寿险",
    "健康医疗险",
    "机动车商业险",
)
CUSTOMER_CLASSES = ("重点客户", "普通客户", "渠道客户", "战略客户")
BUSINESS_TYPES = ("新增", "续保", "增购", "批改续期")
BOOLEAN_YES = ("是", "Y", "TRUE", "有效")

PROPERTY_PRODUCTS = (
    "企业财产险",
    "雇主责任险",
    "公众责任险",
    "建筑工程一切险",
    "机动车商业险",
    "货物运输险",
    "安全生产责任险",
    "机械设备损坏险",
)
LIFE_PRODUCTS = (
    "终身寿险",
    "定期寿险",
    "年金保险",
    "两全保险",
    "重大疾病保险",
    "医疗保险",
    "团体寿险",
    "长期护理保险",
)
PROPERTY_PRIMARY_CATEGORIES = ("产险", "责任险", "车险", "企财险", "工程险")
LIFE_PRIMARY_CATEGORIES = ("寿险", "健康险", "年金险", "团险")
PROPERTY_SECONDARY_CATEGORIES = ("直销", "渠道", "经代", "续保", "团单")
LIFE_SECONDARY_CATEGORIES = ("个险", "团险", "银保", "续期", "健康管理")
PROPERTY_DEPARTMENTS = (
    "产险事业群直营部",
    "产险事业群渠道部",
    "责任险项目部",
    "车险发展部",
    "企财险团队",
)
LIFE_DEPARTMENTS = (
    "寿险事业群个险部",
    "寿险事业群团险部",
    "寿险事业群银保部",
    "健康险产品部",
    "年金保险服务部",
)
OPS_DEPARTMENTS = (
    "经营分析组",
    "主数据治理组",
    "中台运营组",
    "指标管理组",
    "配置维护组",
)

WORKBOOK_THEME_CONFIG = {
    "property_2025": {
        "division_name": "澄岳保险集团产险事业群",
        "sheet_map": {
            "文旅产业链+旅责险+责任保险": "综合产险业务台账",
            "旅游意外险团险": "团体短期险台账",
            "咨询费": "风险管理服务收入",
            "分入明细": "分入业务台账",
            "线下分入明细": "线下分入业务",
        },
        "sheet_title_map": {
            "文旅产业链+旅责险+责任保险": "澄岳保险集团产险事业群综合产险业务台账",
            "旅游意外险团险": "澄岳保险集团产险事业群团体短期险业务台账",
            "咨询费": "澄岳保险集团产险事业群风险管理服务收入台账",
            "分入明细": "澄岳保险集团产险事业群分入业务台账",
            "线下分入明细": "澄岳保险集团产险事业群线下分入业务台账",
        },
    },
    "life_2026": {
        "division_name": "澄岳保险集团寿险事业群",
        "sheet_map": {
            "旅责险": "个险长期险台账",
            "团意险": "团险保障台账",
            "咨询费": "健康服务收入",
            "总部分入-线上": "寿险分入业务台账",
        },
        "sheet_title_map": {
            "旅责险": "澄岳保险集团寿险事业群个险长期险业务台账",
            "团意险": "澄岳保险集团寿险事业群团险保障业务台账",
            "咨询费": "澄岳保险集团寿险事业群健康服务收入台账",
            "总部分入-线上": "澄岳保险集团寿险事业群分入业务台账",
        },
    },
    "ops_center": {
        "division_name": "澄岳保险集团经营管理中台",
        "sheet_map": {
            "首页": "中台首页",
            "系统参数": "中台参数",
            "基础数据元信息": "数据资产说明",
            "Sheet映射": "业务映射配置",
            "字段映射": "字段标准映射",
            "险种映射": "产品分类映射",
            "员工信息": "人员主数据",
            "客户信息": "客户主数据",
            "指标维护": "指标配置",
            "新增客户规则": "新客规则配置",
            "运行日志": "中台运行日志",
            "使用说明": "操作说明",
        },
        "sheet_title_map": {
            "首页": "澄岳保险集团经营管理中台首页",
            "系统参数": "澄岳保险集团经营管理中台参数配置",
            "基础数据元信息": "澄岳保险集团经营管理中台数据资产说明",
            "Sheet映射": "澄岳保险集团经营管理中台业务映射配置",
            "字段映射": "澄岳保险集团经营管理中台字段标准映射",
            "险种映射": "澄岳保险集团经营管理中台产品分类映射",
            "员工信息": "澄岳保险集团经营管理中台人员主数据",
            "客户信息": "澄岳保险集团经营管理中台客户主数据",
            "指标维护": "澄岳保险集团经营管理中台指标配置",
            "新增客户规则": "澄岳保险集团经营管理中台新客规则配置",
            "运行日志": "澄岳保险集团经营管理中台运行日志",
            "使用说明": "澄岳保险集团经营管理中台操作说明",
        },
    },
}


# 2026-03-22 13:46 CST: 修改原因和目的：把单行生成时反复复用的业务上下文收敛成结构体，保证一个 row 内字段彼此一致。
@dataclass(frozen=True)
class RowContext:
    row_index: int
    month: int
    year: int
    month_label: str
    customer_name: str
    standard_customer_name: str
    insurer_name: str
    employee_name: str
    backup_employee_name: str
    region: str
    department: str
    business_channel: str
    primary_category: str
    secondary_category: str
    insurance_product: str
    business_type: str
    customer_class: str
    contract_code: str
    policy_code: str
    phone_number: str
    id_code: str
    credit_code: str
    amount_income: float
    amount_premium: float
    amount_insured: float
    ratio_percent: float
    created_date: date
    effective_date: date
    expiry_date: date


# 2026-03-22 13:46 CST: 修改原因和目的：用统一入口处理金额格式，避免不同列写出过多小数导致假数据不像业务台账。
def round_money(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


# 2026-03-22 13:46 CST: 修改原因和目的：把淡旺季规则显式化，保证假数据具备“淡季低、旺季高”的业务节奏。
def season_factor(month: int) -> float:
    factors = {
        1: 0.72,
        2: 0.58,
        3: 0.81,
        4: 0.92,
        5: 1.03,
        6: 1.18,
        7: 1.42,
        8: 1.36,
        9: 1.28,
        10: 0.96,
        11: 1.21,
        12: 1.33,
    }
    return factors.get(month, 1.0)


# 2026-03-22 13:46 CST: 修改原因和目的：让不同工作簿维持不同年份语境，避免 2025 文件里混入 2026 月份。
def resolve_year(workbook_kind: str) -> int:
    if workbook_kind in {"income_2025", "property_2025"}:
        return 2025
    return 2026


# 2026-03-22 13:46 CST: 修改原因和目的：集中封装确定性取值，既能去敏，又能让重复运行得到稳定结果便于验证。
def pick(values: Sequence[str], index: int, offset: int = 0) -> str:
    return values[(index + offset) % len(values)]


# 2026-03-22 15:32 CST: 修改原因和目的：把产险、寿险、中台三类主题的数据词库拆开，确保假数据语义与工作簿主题一致。
def resolve_theme_profile(workbook_kind: str) -> dict[str, Sequence[str]]:
    if workbook_kind == "property_2025":
        return {
            "departments": PROPERTY_DEPARTMENTS,
            "primary_categories": PROPERTY_PRIMARY_CATEGORIES,
            "secondary_categories": PROPERTY_SECONDARY_CATEGORIES,
            "products": PROPERTY_PRODUCTS,
        }
    if workbook_kind == "life_2026":
        return {
            "departments": LIFE_DEPARTMENTS,
            "primary_categories": LIFE_PRIMARY_CATEGORIES,
            "secondary_categories": LIFE_SECONDARY_CATEGORIES,
            "products": LIFE_PRODUCTS,
        }
    if workbook_kind == "ops_center":
        return {
            "departments": OPS_DEPARTMENTS,
            "primary_categories": ("经营管理", "主数据", "指标管理", "流程配置"),
            "secondary_categories": ("中台", "配置", "治理", "运营"),
            "products": ("客户主数据", "指标配置", "映射规则", "运行日志"),
        }
    return {
        "departments": DEPARTMENTS,
        "primary_categories": PRIMARY_CATEGORIES,
        "secondary_categories": SECONDARY_CATEGORIES,
        "products": INSURANCE_PRODUCTS,
    }


# 2026-03-22 13:46 CST: 修改原因和目的：按行生成完整业务上下文，后续各列映射只消费上下文，减少重复逻辑并保持 SRP。
def build_row_context(row_index: int, workbook_kind: str, sheet_name: str) -> RowContext:
    month = (row_index % 12) + 1
    year = resolve_year(workbook_kind)
    theme_profile = resolve_theme_profile(workbook_kind)
    region = pick(REGIONS, row_index, len(sheet_name))
    customer_name = f"{region}{pick(CUSTOMER_PREFIXES, row_index)}{pick(CUSTOMER_SUFFIXES, row_index, 2)}"
    insurer_name = f"{pick(INSURER_PREFIXES, row_index, 1)}{pick(INSURER_SUFFIXES, row_index, 3)}"
    employee_name = pick(PERSON_NAMES, row_index, 4)
    backup_employee_name = pick(PERSON_NAMES, row_index, 7)
    department = pick(theme_profile["departments"], row_index, len(workbook_kind))
    business_channel = pick(BUSINESS_CHANNELS, row_index, month)
    primary_category = pick(theme_profile["primary_categories"], row_index, month // 2)
    secondary_category = pick(theme_profile["secondary_categories"], row_index, len(sheet_name) + month)
    insurance_product = pick(theme_profile["products"], row_index, month)
    business_type = pick(BUSINESS_TYPES, row_index, month + 1)
    customer_class = pick(CUSTOMER_CLASSES, row_index, month + 2)

    base_income = 18000 + (row_index % 17) * 1450 + len(sheet_name) * 320
    amount_income = round_money(base_income * season_factor(month))
    amount_premium = round_money(amount_income * (1.65 + (row_index % 5) * 0.18))
    amount_insured = round_money(amount_premium * (60 + (row_index % 9) * 8))
    ratio_percent = round_money(8 + (row_index % 7) * 4.5)

    created_date = date(year, month, min(25, 5 + (row_index % 20)))
    effective_date = created_date + timedelta(days=1 + (row_index % 5))
    expiry_date = effective_date + timedelta(days=365)

    contract_code = f"T{str(year)[2:]}{month:02d}{(row_index % 28) + 1:02d}{1000000000 + row_index:010d}"
    policy_code = f"P{str(year)[2:]}{month:02d}{(row_index % 28) + 1:02d}{2000000000 + row_index:010d}"
    phone_number = f"13{(row_index % 9) + 1}{80000000 + row_index:08d}"[:11]
    id_code = f"{650100 + row_index % 900:06d}{year - 30:04d}{month:02d}{(row_index % 28) + 1:02d}{1000 + row_index % 8999:04d}"
    credit_code = f"91{650100 + row_index % 800:06d}{300000 + row_index % 700000:06d}{(row_index % 35):02d}"[:18]

    return RowContext(
        row_index=row_index,
        month=month,
        year=year,
        month_label=f"{year}-{month:02d}",
        customer_name=customer_name,
        standard_customer_name=customer_name,
        insurer_name=insurer_name,
        employee_name=employee_name,
        backup_employee_name=backup_employee_name,
        region=region,
        department=department,
        business_channel=business_channel,
        primary_category=primary_category,
        secondary_category=secondary_category,
        insurance_product=insurance_product,
        business_type=business_type,
        customer_class=customer_class,
        contract_code=contract_code,
        policy_code=policy_code,
        phone_number=phone_number,
        id_code=id_code,
        credit_code=credit_code,
        amount_income=amount_income,
        amount_premium=amount_premium,
        amount_insured=amount_insured,
        ratio_percent=ratio_percent,
        created_date=created_date,
        effective_date=effective_date,
        expiry_date=expiry_date,
    )


# 2026-03-22 13:46 CST: 修改原因和目的：把表头匹配逻辑封装起来，后续扩展关键字时只改一处即可。
def header_matches(header: str, keywords: Iterable[str]) -> bool:
    lowered = header.strip().lower()
    return any(keyword.lower() in lowered for keyword in keywords)


# 2026-03-22 15:32 CST: 修改原因和目的：将首行标题、列头别名和 sheet 改名统一到主题配置里，避免分散判断导致命名不一致。
def resolve_theme_config(workbook_kind: str) -> dict:
    return WORKBOOK_THEME_CONFIG.get(workbook_kind, {"division_name": "", "sheet_map": {}, "sheet_title_map": {}})


# 2026-03-22 15:36 CST: 修改原因和目的：保留旧 workbook_kind 的原始行为，只让新引入的寿险/产险/中台主题执行元数据重构。
def is_themed_workbook_kind(workbook_kind: str) -> bool:
    return workbook_kind in WORKBOOK_THEME_CONFIG


# 2026-03-22 15:32 CST: 修改原因和目的：把列头脱敏别名统一封装，确保“第一行标题/表头说明”能整体切换为寿险、产险或中台语境。
def alias_header_text(value, workbook_kind: str):
    if value in (None, ""):
        return value

    text = str(value).strip()
    common_aliases = {
        "经营单位": "经营机构",
        "业务部门": "所属团队",
        "委托合同编号": "业务编号",
        "客户名称": "投保主体" if workbook_kind != "ops_center" else "主体名称",
        "标准客户名称": "标准主体名称",
        "一级业务类别": "一级板块",
        "二级业务类别": "二级板块",
        "客户": "投保主体",
        "项目": "业务方案" if workbook_kind != "ops_center" else "配置项",
        "业务分类": "经营渠道",
        "险种": "标准产品分类" if workbook_kind != "ops_center" else "产品大类",
        "业务人员": "业务经理",
        "会计月度": "会计期间",
        "业务收入（人民币）": "经营收入（元）",
        "报表收入": "经营收入（元）",
        "分入单位（填报单位）": "填报机构",
        "分出单位": "对手机构",
        "值": "当前值",
        "内部参数名": "参数键",
        "参数值": "参数值",
        "参数用途": "参数用途",
        "内部映射键": "映射键",
        "Sheet名称": "工作表名称",
        "源Sheet": "源工作表",
        "源字段名": "源字段",
        "标准字段名": "标准字段",
        "源险种": "源产品分类",
        "标准险种": "标准产品分类",
        "员工姓名": "人员姓名",
        "角色类型": "岗位类型",
        "年初客户家数": "期初客户数",
        "客户分类": "主体分层",
        "A角": "主责经理",
        "B角": "协同经理",
        "员工": "人员",
        "年份": "年度",
        "输入台账": "输入台账",
        "最终成品文件": "输出文件",
    }
    workbook_aliases = {
        "property_2025": {
            "业务方案": "产险方案",
            "标准产品分类": "产险产品分类",
        },
        "life_2026": {
            "业务方案": "寿险方案",
            "标准产品分类": "寿险产品分类",
        },
        "ops_center": {
            "业务方案": "配置项",
            "所属团队": "所属模块",
            "经营收入（元）": "指标值",
        },
    }

    aliased = common_aliases.get(text, text)
    aliased = workbook_aliases.get(workbook_kind, {}).get(aliased, aliased)
    return aliased


# 2026-03-22 15:32 CST: 修改原因和目的：统一生成每个 sheet 的展示名称，保证文件内部命名与目标主题一致。
def themed_sheet_name(original_sheet_name: str, workbook_kind: str) -> str:
    theme_config = resolve_theme_config(workbook_kind)
    return theme_config["sheet_map"].get(original_sheet_name, original_sheet_name)


# 2026-03-22 15:32 CST: 修改原因和目的：当首行是说明标题而不是字段行时，给出新的主题标题，避免仍暴露原始业务背景。
def themed_sheet_title(original_sheet_name: str, workbook_kind: str) -> str:
    theme_config = resolve_theme_config(workbook_kind)
    return theme_config["sheet_title_map"].get(
        original_sheet_name,
        f"{theme_config.get('division_name', '澄岳保险集团')} - {themed_sheet_name(original_sheet_name, workbook_kind)}",
    )


# 2026-03-22 13:46 CST: 修改原因和目的：对未知列提供按原值类型兜底的假值，避免因为表头命名差异导致整列空白。
def fallback_by_original_value(context: RowContext, original_value):
    if original_value is None:
        return f"样例值-{context.row_index + 1}"
    if isinstance(original_value, bool):
        return True
    if isinstance(original_value, int) and not isinstance(original_value, bool):
        return int(context.amount_income)
    if isinstance(original_value, float):
        return context.amount_income
    if isinstance(original_value, datetime):
        return datetime.combine(context.created_date, datetime.min.time())
    if isinstance(original_value, date):
        return context.created_date
    return f"{context.region}样例字段{context.row_index + 1}"


# 2026-03-22 14:26 CST: 修改原因和目的：COM 写 Excel 日期单元格时不能直接吃 `date`，这里统一转成 `datetime`，避免真实文件批量写入中断。
def normalize_value_for_com(value):
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, datetime.min.time())
    return value


# 2026-03-22 13:46 CST: 修改原因和目的：对常见业务字段做语义映射，让每一列像真数据而不是简单随机串。
def generate_cell_value(header: str, context: RowContext, original_value=None):
    cleaned = (header or "").strip()

    if not cleaned:
        return fallback_by_original_value(context, original_value)
    if header_matches(cleaned, ("客户名称", "投保主体", "主体名称")) and "标准" not in cleaned:
        return context.customer_name
    if header_matches(cleaned, ("标准客户名称", "标准主体名称")):
        return context.standard_customer_name
    if header_matches(cleaned, ("客户分类", "主体分层")):
        return context.customer_class
    if header_matches(cleaned, ("保险公司", "承保公司", "保险人", "承保机构")):
        return context.insurer_name
    if header_matches(cleaned, ("经营单位", "分入单位", "分出单位", "单位", "机构")):
        return f"{context.region}{pick(('分公司', '营业部', '中心支公司', '营销服务部'), context.row_index)}"
    if header_matches(cleaned, ("业务部门", "部门", "所属团队", "所属模块")):
        return context.department
    if header_matches(cleaned, ("业务人员", "经办人", "客户经理", "A角", "负责人", "员工姓名", "员工", "业务经理", "主责经理")):
        return context.employee_name
    if header_matches(cleaned, ("B角", "协办人", "协同经理")):
        return context.backup_employee_name
    if header_matches(cleaned, ("联系电话", "电话", "手机号", "手机")):
        return context.phone_number
    if header_matches(cleaned, ("身份证", "证件号")):
        return context.id_code
    if header_matches(cleaned, ("统一社会信用代码", "信用代码")):
        return context.credit_code
    if header_matches(cleaned, ("委托合同编号", "合同编号", "合同号", "单号", "业务编号")):
        return context.contract_code
    if header_matches(cleaned, ("保单号", "投保单号", "保单编号")):
        return context.policy_code
    if header_matches(cleaned, ("一级业务类别", "一级板块")):
        return context.primary_category
    if header_matches(cleaned, ("二级业务类别", "二级板块")):
        return context.secondary_category
    if header_matches(cleaned, ("业务分类", "渠道")):
        return context.business_channel
    if header_matches(cleaned, ("险种", "产品", "产品分类")):
        return context.insurance_product
    if header_matches(cleaned, ("项目", "配置项", "方案")):
        return f"{context.insurance_product}年度保障计划"
    if header_matches(cleaned, ("新增/续保", "业务类型")):
        return context.business_type
    if header_matches(cleaned, ("会计月度", "月份", "会计期间")):
        return context.month_label
    if header_matches(cleaned, ("年份", "年度")):
        return context.year
    if header_matches(cleaned, ("日期", "时间", "出单日", "起保日", "生效日")):
        if header_matches(cleaned, ("终保", "到期")):
            return context.expiry_date
        if header_matches(cleaned, ("起保", "生效")):
            return context.effective_date
        return context.created_date
    if header_matches(cleaned, ("客户", "主体")) and "名称" not in cleaned:
        return context.customer_name
    if header_matches(cleaned, ("地区", "城市", "区域")):
        return context.region
    if header_matches(cleaned, ("序号",)):
        return context.row_index + 1
    if header_matches(cleaned, ("是否", "启用", "有效")):
        return "是"
    if header_matches(cleaned, ("排序", "优先级")):
        return (context.row_index % 999) + 1
    if header_matches(cleaned, ("比例", "占比", "费率", "折扣")):
        if "%" in str(original_value):
            return f"{context.ratio_percent}%"
        return round_money(context.ratio_percent / 100)
    if header_matches(cleaned, ("保额", "限额")):
        return context.amount_insured
    if header_matches(cleaned, ("保费", "收入", "金额", "费用", "佣金", "报表收入", "目标", "预算", "赔款", "指标值")):
        if "保费" in cleaned:
            return context.amount_premium
        if "赔款" in cleaned:
            return round_money(context.amount_income * 0.18)
        if "佣金" in cleaned:
            return round_money(context.amount_income * 0.12)
        return context.amount_income
    if header_matches(cleaned, ("备注", "说明")):
        return f"2026-03-22脱敏生成：保留结构并替换为虚构保险业务样本"
    return fallback_by_original_value(context, original_value)


# 2026-03-22 14:38 CST: 修改原因和目的：把表头解析编译成列级 resolver，真实大文件运行时只解析一次列语义，避免每个单元格重复做字符串匹配。
def build_column_resolver(header: str, sample_value=None):
    cleaned = (header or "").strip()

    if not cleaned:
        return lambda context: fallback_by_original_value(context, sample_value)
    if header_matches(cleaned, ("客户名称", "投保主体", "主体名称")) and "标准" not in cleaned:
        return lambda context: context.customer_name
    if header_matches(cleaned, ("标准客户名称", "标准主体名称")):
        return lambda context: context.standard_customer_name
    if header_matches(cleaned, ("客户分类", "主体分层")):
        return lambda context: context.customer_class
    if header_matches(cleaned, ("保险公司", "承保公司", "保险人", "承保机构")):
        return lambda context: context.insurer_name
    if header_matches(cleaned, ("经营单位", "分入单位", "分出单位", "单位", "机构")):
        return lambda context: f"{context.region}{pick(('分公司', '营业部', '中心支公司', '营销服务部'), context.row_index)}"
    if header_matches(cleaned, ("业务部门", "部门", "所属团队", "所属模块")):
        return lambda context: context.department
    if header_matches(cleaned, ("业务人员", "经办人", "客户经理", "A角", "负责人", "员工姓名", "员工", "业务经理", "主责经理")):
        return lambda context: context.employee_name
    if header_matches(cleaned, ("B角", "协办人", "协同经理")):
        return lambda context: context.backup_employee_name
    if header_matches(cleaned, ("联系电话", "电话", "手机号", "手机")):
        return lambda context: context.phone_number
    if header_matches(cleaned, ("身份证", "证件号")):
        return lambda context: context.id_code
    if header_matches(cleaned, ("统一社会信用代码", "信用代码")):
        return lambda context: context.credit_code
    if header_matches(cleaned, ("委托合同编号", "合同编号", "合同号", "单号", "业务编号")):
        return lambda context: context.contract_code
    if header_matches(cleaned, ("保单号", "投保单号", "保单编号")):
        return lambda context: context.policy_code
    if header_matches(cleaned, ("一级业务类别", "一级板块")):
        return lambda context: context.primary_category
    if header_matches(cleaned, ("二级业务类别", "二级板块")):
        return lambda context: context.secondary_category
    if header_matches(cleaned, ("业务分类", "渠道")):
        return lambda context: context.business_channel
    if header_matches(cleaned, ("险种", "产品", "产品分类")):
        return lambda context: context.insurance_product
    if header_matches(cleaned, ("项目", "配置项", "方案")):
        return lambda context: f"{context.insurance_product}年度保障计划"
    if header_matches(cleaned, ("新增/续保", "业务类型")):
        return lambda context: context.business_type
    if header_matches(cleaned, ("会计月度", "月份", "会计期间")):
        return lambda context: context.month_label
    if header_matches(cleaned, ("年份", "年度")):
        return lambda context: context.year
    if header_matches(cleaned, ("日期", "时间", "出单日", "起保日", "生效日")):
        if header_matches(cleaned, ("终保", "到期")):
            return lambda context: context.expiry_date
        if header_matches(cleaned, ("起保", "生效")):
            return lambda context: context.effective_date
        return lambda context: context.created_date
    if header_matches(cleaned, ("客户", "主体")) and "名称" not in cleaned:
        return lambda context: context.customer_name
    if header_matches(cleaned, ("地区", "城市", "区域")):
        return lambda context: context.region
    if header_matches(cleaned, ("序号",)):
        return lambda context: context.row_index + 1
    if header_matches(cleaned, ("是否", "启用", "有效")):
        return lambda context: "是"
    if header_matches(cleaned, ("排序", "优先级")):
        return lambda context: (context.row_index % 999) + 1
    if header_matches(cleaned, ("比例", "占比", "费率", "折扣")):
        if "%" in str(sample_value):
            return lambda context: f"{context.ratio_percent}%"
        return lambda context: round_money(context.ratio_percent / 100)
    if header_matches(cleaned, ("保额", "限额")):
        return lambda context: context.amount_insured
    if header_matches(cleaned, ("保费", "收入", "金额", "费用", "佣金", "报表收入", "目标", "预算", "赔款", "指标值")):
        if "保费" in cleaned:
            return lambda context: context.amount_premium
        if "赔款" in cleaned:
            return lambda context: round_money(context.amount_income * 0.18)
        if "佣金" in cleaned:
            return lambda context: round_money(context.amount_income * 0.12)
        return lambda context: context.amount_income
    if header_matches(cleaned, ("备注", "说明")):
        return lambda context: "2026-03-22脱敏生成：保留结构并替换为虚构保险业务样本"
    return lambda context: fallback_by_original_value(context, sample_value)


# 2026-03-22 14:38 CST: 修改原因和目的：提供纯函数接口给测试直接验证季节性和字段逻辑，同时复用列 resolver 提升真实运行性能。
def build_fake_rows(
    headers: Sequence[str],
    row_count: int,
    workbook_kind: str,
    sheet_name: str,
    sample_values: Sequence | None = None,
):
    if sample_values is None:
        sample_values = [None] * len(headers)
    resolvers = [build_column_resolver(header, sample_values[index]) for index, header in enumerate(headers)]
    rows = []
    for index in range(row_count):
        context = build_row_context(index, workbook_kind, sheet_name)
        row = [resolver(context) for resolver in resolvers]
        rows.append(row)
    return rows


# 2026-03-22 13:46 CST: 修改原因和目的：表头行不一定永远是第 1 行，这里做保守识别，尽量保留顶部说明和结构。
def detect_header_row(worksheet, scan_limit: int = 12) -> int:
    best_row = 1
    best_score = -1
    for row_index in range(1, min(scan_limit, worksheet.max_row) + 1):
        values = [worksheet.cell(row=row_index, column=column).value for column in range(1, worksheet.max_column + 1)]
        non_empty = sum(1 for value in values if value not in (None, ""))
        text_like = sum(1 for value in values if isinstance(value, str) and value.strip())
        score = non_empty * 3 + text_like
        if score > best_score:
            best_row = row_index
            best_score = score
    return best_row


# 2026-03-22 13:46 CST: 修改原因和目的：从底部回扫最后一行，减少 max_row 因格式残留而虚高时对数据区判断的干扰。
def detect_last_data_row(worksheet, header_row: int) -> int:
    for row_index in range(worksheet.max_row, header_row, -1):
        for column in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=row_index, column=column).value not in (None, ""):
                return row_index
    return header_row


# 2026-03-22 15:32 CST: 修改原因和目的：统一改写首行标题或列头，并在需要时同步改写真正的表头行，满足“sheet 名和第一行标题都要脱敏”。
def rewrite_sheet_metadata_openpyxl(worksheet, workbook_kind: str, original_sheet_name: str, header_row: int) -> None:
    if not is_themed_workbook_kind(workbook_kind):
        return
    first_row_has_multiple_cells = sum(
        1 for column in range(1, worksheet.max_column + 1) if worksheet.cell(row=1, column=column).value not in (None, "")
    ) > 1

    if header_row == 1 or first_row_has_multiple_cells:
        for column in range(1, worksheet.max_column + 1):
            worksheet.cell(row=1, column=column).value = alias_header_text(worksheet.cell(row=1, column=column).value, workbook_kind)
    else:
        worksheet.cell(row=1, column=1).value = themed_sheet_title(original_sheet_name, workbook_kind)

    if header_row > 1:
        for column in range(1, worksheet.max_column + 1):
            worksheet.cell(row=header_row, column=column).value = alias_header_text(
                worksheet.cell(row=header_row, column=column).value,
                workbook_kind,
            )

    worksheet.title = themed_sheet_name(original_sheet_name, workbook_kind)


# 2026-03-22 15:32 CST: 修改原因和目的：给 COM 路径补齐与 openpyxl 一致的元数据重写，保证真实大文件与测试样本行为一致。
def rewrite_sheet_metadata_com(worksheet, workbook_kind: str, original_sheet_name: str, header_row: int, first_col: int, column_count: int) -> None:
    if not is_themed_workbook_kind(workbook_kind):
        return
    first_row_non_empty = 0
    for column_index in range(first_col, first_col + column_count):
        if worksheet.Cells(1, column_index).Value not in (None, ""):
            first_row_non_empty += 1

    if header_row == 1 or first_row_non_empty > 1:
        for column_index in range(first_col, first_col + column_count):
            worksheet.Cells(1, column_index).Value = alias_header_text(worksheet.Cells(1, column_index).Value, workbook_kind)
    else:
        worksheet.Cells(1, first_col).Value = themed_sheet_title(original_sheet_name, workbook_kind)

    if header_row > 1:
        for column_index in range(first_col, first_col + column_count):
            worksheet.Cells(header_row, column_index).Value = alias_header_text(
                worksheet.Cells(header_row, column_index).Value,
                workbook_kind,
            )

    worksheet.Name = themed_sheet_name(original_sheet_name, workbook_kind)


# 2026-03-22 13:46 CST: 修改原因和目的：统一输出命名规则，真实运行时避免覆盖原文件，也方便用户一眼识别脱敏副本。
def build_output_path(source_path: Path, output_dir: Path, output_name: str | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    if output_name:
        candidate = output_dir / output_name
    else:
        candidate = output_dir / f"{source_path.stem}-脱敏{source_path.suffix}"

    # 2026-03-22 15:02 CST: 修改原因和目的：如果同名脱敏文件已存在或被占用，自动改成带时间戳的新文件名，避免覆盖失败中断整批处理。
    if candidate.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate = candidate.with_name(f"{candidate.stem}-{stamp}{candidate.suffix}")
    return candidate


# 2026-03-22 14:08 CST: 修改原因和目的：显式探测 Excel COM 是否可用，真实大文件优先走更快的桌面 Excel 写入路径。
def excel_com_is_available() -> bool:
    return pythoncom is not None and win32 is not None


# 2026-03-22 14:08 CST: 修改原因和目的：在 COM 路径下复用表头识别能力，减少真实文件中标题区被误判为数据区的风险。
def detect_header_row_com(worksheet, first_row: int, first_col: int, row_count: int, column_count: int, scan_limit: int = 12) -> int:
    best_row = first_row
    best_score = -1
    last_scan_row = min(first_row + row_count - 1, first_row + scan_limit - 1)
    for row_index in range(first_row, last_scan_row + 1):
        non_empty = 0
        text_like = 0
        for column_index in range(first_col, first_col + column_count):
            value = worksheet.Cells(row_index, column_index).Value
            if value not in (None, ""):
                non_empty += 1
            if isinstance(value, str) and value.strip():
                text_like += 1
        score = non_empty * 3 + text_like
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


# 2026-03-22 13:46 CST: 修改原因和目的：把“复制副本 + 重写目标sheet”收敛成单一入口，便于测试、复用和真实任务执行。
def rewrite_workbook_copy(
    source_path: Path,
    output_dir: Path,
    target_sheet_names: set[str] | None,
    workbook_kind: str,
    output_name: str | None = None,
) -> Path:
    output_path = build_output_path(source_path, output_dir, output_name)
    shutil.copy2(source_path, output_path)

    keep_vba = output_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(output_path, keep_vba=keep_vba)
    try:
        sheets = workbook.worksheets
        target_names = {sheet.title for sheet in sheets} if target_sheet_names is None else set(target_sheet_names)

        for worksheet in sheets:
            original_sheet_name = worksheet.title
            header_row = detect_header_row(worksheet)
            rewrite_sheet_metadata_openpyxl(worksheet, workbook_kind, original_sheet_name, header_row)

            if original_sheet_name not in target_names:
                continue

            last_data_row = detect_last_data_row(worksheet, header_row)
            if last_data_row <= header_row:
                continue

            headers = [worksheet.cell(row=header_row, column=column).value or "" for column in range(1, worksheet.max_column + 1)]
            row_count = last_data_row - header_row
            sample_values = [worksheet.cell(row=header_row + 1, column=column).value for column in range(1, worksheet.max_column + 1)]
            fake_rows = build_fake_rows(
                headers,
                row_count,
                workbook_kind,
                themed_sheet_name(original_sheet_name, workbook_kind),
                sample_values,
            )

            for offset, fake_row in enumerate(fake_rows):
                row_index = header_row + 1 + offset
                for column, fake_value in enumerate(fake_row, start=1):
                    worksheet.cell(row=row_index, column=column).value = fake_value

        workbook.save(output_path)
    finally:
        # 2026-03-22 13:49 CST: 修改原因和目的：显式关闭 openpyxl 在 keep_vba 模式下持有的压缩包句柄，避免测试和真实运行结束时残留 ZipFile 告警。
        if getattr(workbook, "vba_archive", None) is not None:
            workbook.vba_archive.close()
        if getattr(workbook, "_archive", None) is not None:
            workbook._archive.close()
        workbook.close()
    return output_path


# 2026-03-22 14:08 CST: 修改原因和目的：对真实大工作簿提供 COM 批量写入路径，避免 openpyxl 在十万级行数下执行过慢。
def rewrite_workbook_copy_with_excel_com(
    source_path: Path,
    output_dir: Path,
    target_sheet_names: set[str] | None,
    workbook_kind: str,
    output_name: str | None = None,
) -> Path:
    if not excel_com_is_available():
        return rewrite_workbook_copy(source_path, output_dir, target_sheet_names, workbook_kind, output_name)

    output_path = build_output_path(source_path, output_dir, output_name)
    shutil.copy2(source_path, output_path)

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(str(output_path))
    try:
        worksheet_names = {str(sheet.Name) for sheet in workbook.Worksheets}
        target_names = worksheet_names if target_sheet_names is None else set(target_sheet_names)

        for worksheet in workbook.Worksheets:
            original_sheet_name = str(worksheet.Name)

            used_range = worksheet.UsedRange
            first_row = int(used_range.Row)
            first_col = int(used_range.Column)
            row_count = int(used_range.Rows.Count)
            column_count = int(used_range.Columns.Count)

            if row_count <= 1 or column_count <= 0:
                rewrite_sheet_metadata_com(worksheet, workbook_kind, original_sheet_name, 1, first_col, column_count)
                continue

            header_row = detect_header_row_com(worksheet, first_row, first_col, row_count, column_count)
            rewrite_sheet_metadata_com(worksheet, workbook_kind, original_sheet_name, header_row, first_col, column_count)

            if original_sheet_name not in target_names:
                continue

            data_row_count = first_row + row_count - header_row - 1
            if data_row_count <= 0:
                continue

            headers = []
            sample_values = []
            for column_index in range(first_col, first_col + column_count):
                headers.append(str(worksheet.Cells(header_row, column_index).Value or ""))
                sample_values.append(worksheet.Cells(header_row + 1, column_index).Value)

            matrix = []
            fake_rows = build_fake_rows(headers, data_row_count, workbook_kind, themed_sheet_name(original_sheet_name, workbook_kind), sample_values)
            for fake_row in fake_rows:
                matrix.append(tuple(normalize_value_for_com(value) for value in fake_row))

            target_range = worksheet.Range(
                worksheet.Cells(header_row + 1, first_col),
                worksheet.Cells(header_row + data_row_count, first_col + column_count - 1),
            )
            target_range.Value = tuple(matrix)

        workbook.Save()
    finally:
        workbook.Close(True)
        excel.Quit()
        pythoncom.CoUninitialize()
    return output_path


# 2026-03-22 13:46 CST: 修改原因和目的：批量执行默认任务并打印结果，方便这次直接交付，也方便后续重复运行。
def run_default_jobs(output_dir: Path) -> list[Path]:
    outputs = []
    for job in DEFAULT_JOBS:
        if excel_com_is_available():
            output = rewrite_workbook_copy_with_excel_com(
                source_path=job["source_path"],
                output_dir=output_dir,
                target_sheet_names=job["target_sheet_names"],
                workbook_kind=job["workbook_kind"],
                output_name=job["output_name"],
            )
        else:
            output = rewrite_workbook_copy(
                source_path=job["source_path"],
                output_dir=output_dir,
                target_sheet_names=job["target_sheet_names"],
                workbook_kind=job["workbook_kind"],
                output_name=job["output_name"],
            )
        outputs.append(output)
    return outputs


# 2026-03-22 13:46 CST: 修改原因和目的：提供命令行入口，既能直接跑默认客户任务，也能在本地做单次复用。
def main() -> int:
    parser = argparse.ArgumentParser(description="复制 Excel 并把目标 sheet 的业务数据替换成假数据。")
    parser.add_argument(
        "--output-dir",
        default=r"D:\Excel测试\脱敏数据",
        help="脱敏副本输出目录，默认使用本次任务要求的目录。",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    outputs = run_default_jobs(output_dir)
    for output in outputs:
        print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
